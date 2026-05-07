from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from _common.paths import FASE5_OUTPUTS, MVP_ROOT
from FASE5.src.validate import EXPECTED_EXCEL_SHEETS


EXPECTED_OUTPUTS = {
    "feature_matrix_mvp.csv",
    "coverage_report_mvp.csv",
    "mvp_countries.csv",
    "mvp_variables_catalog.csv",
    "mvp_transform_params.csv",
    "mvp_train_test_split.csv",
    "MVP_AUDITABLE.xlsx",
    "fase5_manifest.json",
}


def test_phase5_outputs_exist_and_are_non_empty(phase5_results):
    for filename in EXPECTED_OUTPUTS:
        path = FASE5_OUTPUTS / filename
        assert path.exists(), filename
        assert path.stat().st_size > 0, filename


def test_feature_matrix_and_split_contracts(phase5_results):
    feature_matrix = pd.read_csv(FASE5_OUTPUTS / "feature_matrix_mvp.csv")
    split = pd.read_csv(FASE5_OUTPUTS / "mvp_train_test_split.csv")

    assert feature_matrix.shape[0] == 43
    assert feature_matrix["iso3"].is_unique
    assert feature_matrix["split"].isin({"train", "test"}).all()
    assert set(split["split"]) == {"train", "test"}
    assert split.shape[0] == 43


def test_coverage_threshold_contract(phase5_results):
    coverage = pd.read_csv(FASE5_OUTPUTS / "coverage_report_mvp.csv")

    assert coverage.shape[0] == 46
    assert not coverage["below_threshold"].any()
    assert coverage["pct_complete"].ge(30.0).all()


def test_auditable_excel_has_expected_sheets(phase5_results):
    workbook = load_workbook(FASE5_OUTPUTS / "MVP_AUDITABLE.xlsx", read_only=True)

    assert workbook.sheetnames == EXPECTED_EXCEL_SHEETS
    assert workbook["1_Hipotesis"].max_row == 16
    assert workbook["5_Variables_40"].max_row == 47
    assert workbook["5b_Variables_46_Detalle"].max_row == 7
    assert workbook["6_Matriz_40_Humana"].max_row == 44
    assert workbook["6_Matriz_40_Humana"].max_column == 50
    assert workbook["9_Normalizacion"].max_row >= 12
    assert workbook["11_Features_Fase6"].max_row == 44
    assert workbook["11_Features_Fase6"].max_column >= 138
    assert workbook["11b_Features_Fase6_v2"].max_row == 44
    assert workbook["11b_Features_Fase6_v2"].max_column >= 138
    assert workbook["12_Diccionario_Cols"].max_row >= 139


def test_manifest_includes_all_phase5_outputs(phase5_results):
    manifest = json.loads((FASE5_OUTPUTS / "fase5_manifest.json").read_text(encoding="utf-8"))

    for filename in EXPECTED_OUTPUTS - {"fase5_manifest.json"}:
        assert filename in manifest["outputs_hashed"]
    assert manifest["phase5_implementation_hashed"]
    assert (MVP_ROOT / "manifest_mvp.json").exists()
