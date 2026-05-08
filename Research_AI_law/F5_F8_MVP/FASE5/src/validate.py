"""Validacion reproducible de contratos de Fase 5 MVP."""

from __future__ import annotations

import hashlib
import json

import pandas as pd
import yaml
from openpyxl import load_workbook

from _common.paths import FASE5_OUTPUTS, FASE5_PHASE6_READY, MVP_ROOT
from .api import (
    load_coverage_report,
    load_feature_matrix,
    load_manifest,
    load_mvp_countries,
    load_mvp_variables_catalog,
    load_analysis_sample_membership,
    load_transform_params,
)


EXPECTED_OUTPUTS = [
    "feature_matrix_mvp.csv",
    "coverage_report_mvp.csv",
    "mvp_countries.csv",
    "mvp_variables_catalog.csv",
    "mvp_transform_params.csv",
    "analysis_sample_membership.csv",
    "MVP_AUDITABLE.xlsx",
    "fase5_manifest.json",
]

EXPECTED_PHASE6_READY_OUTPUTS = [
    "phase6_feature_matrix.csv",
    "phase6_schema.csv",
    "phase6_schema.json",
    "phase6_column_groups.yaml",
    "phase6_modeling_contract.yaml",
    "phase6_missingness_by_column.csv",
    "phase6_missingness_by_country.csv",
    "phase6_variables_catalog.csv",
    "phase6_transform_params.csv",
    "phase6_analysis_sample_membership.csv",
    "phase6_llm_context.json",
    "phase6_ready_manifest.json",
]

EXPECTED_EXCEL_SHEETS = [
    "0_Leer_Primero",
    "1_Hipotesis",
    "2_Como_Auditar",
    "3_Paises_43",
    "4_Ingreso_Region",
    "5_Variables_40",
    "5b_Variables_46_Detalle",
    "6_Matriz_40_Humana",
    "7_Leyenda_Colores",
    "8_Casos_Atencion",
    "9_Normalizacion",
    "10_Cobertura",
    "11_Features_Fase6",
    "12_Diccionario_Cols",
    "13_Trazabilidad",
    "14_Transformaciones",
    "Muestra_Analitica_v2_1",
]


def _sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()

def validate_no_split_artifacts(outputs_dir: Path) -> None:
    forbidden_files = [
        outputs_dir / "mvp_train_test_split.csv",
        outputs_dir / "phase6_ready" / "phase6_train_test_split.csv",
    ]
    for f in forbidden_files:
        if f.exists():
            raise AssertionError(f"Forbidden split artifact exists: {f}")


def validate_no_split_column(df: pd.DataFrame, name: str) -> None:
    forbidden = {"split", "train", "test", "holdout", "partition"}
    present = forbidden.intersection(set(df.columns))
    if present:
        raise AssertionError(f"Forbidden partition columns in {name}: {sorted(present)}")


def validate_membership(membership: pd.DataFrame) -> None:
    assert len(membership) == 43
    assert membership["iso3"].nunique() == 43
    assert "CHL" in set(membership["iso3"])
    assert membership["is_primary_analysis_sample"].all()
    validate_no_split_column(membership, "analysis_sample_membership")


def validate_transform_policy(transform_params: pd.DataFrame, column_groups: dict) -> None:
    non_estimable = set(
        transform_params.loc[
            transform_params["used_in_primary_modeling"].eq(False),
            "variable_derived"
        ].dropna()
    )
    primary_groups = [
        "X1_regulatory", "X2_control", "Y_Q1_investment", "Y_Q2_adoption",
        "Y_Q3_innovation", "Y_Q5_population_usage", "Y_Q6_public_sector",
        "primary_model_features",
    ]
    for group in primary_groups:
        used = set(column_groups.get(group, []))
        overlap = used & non_estimable
        if overlap:
            raise AssertionError(f"Non-estimable derived features in {group}: {sorted(overlap)}")

def validate_phase5() -> list[str]:
    checks: list[str] = []
    
    validate_no_split_artifacts(FASE5_OUTPUTS)
    
    for filename in EXPECTED_OUTPUTS:
        path = FASE5_OUTPUTS / filename
        if not path.exists() or path.stat().st_size == 0:
            raise AssertionError(f"Output faltante o vacio: {path}")
    for filename in EXPECTED_PHASE6_READY_OUTPUTS:
        path = FASE5_PHASE6_READY / filename
        if not path.exists() or path.stat().st_size == 0:
            raise AssertionError(f"Output tecnico Fase 6 faltante o vacio: {path}")
    checks.append("outputs_exist")

    feature_matrix = load_feature_matrix()
    coverage = load_coverage_report()
    countries = load_mvp_countries()
    variables_catalog = load_mvp_variables_catalog()
    transform_params = load_transform_params()
    membership = load_analysis_sample_membership()
    manifest = load_manifest()

    validate_no_split_column(feature_matrix, "feature_matrix_mvp.csv")
    validate_membership(membership)

    assert feature_matrix.shape[0] == 43
    assert feature_matrix["iso3"].is_unique
    assert countries.shape[0] == 43
    assert variables_catalog.shape[0] == 46
    assert coverage.shape[0] == 46
    assert not coverage["below_threshold"].any()
    assert coverage["pct_complete"].ge(30.0).all()
    assert "log_transform" in set(transform_params["transform_type"])
    assert "robust_zscore" in set(transform_params["transform_type"])
    assert not any(col.endswith("_transform_method") for col in feature_matrix.columns)
    checks.append("data_contracts")

    # The technical bundle intentionally mirrors the canonical Fase 5 feature matrix.
    phase6_feature_matrix = pd.read_csv(FASE5_PHASE6_READY / "phase6_feature_matrix.csv")
    phase6_schema = pd.read_csv(FASE5_PHASE6_READY / "phase6_schema.csv")
    phase6_missing_cols = pd.read_csv(FASE5_PHASE6_READY / "phase6_missingness_by_column.csv")
    phase6_missing_rows = pd.read_csv(FASE5_PHASE6_READY / "phase6_missingness_by_country.csv")
    column_groups = yaml.safe_load((FASE5_PHASE6_READY / "phase6_column_groups.yaml").read_text(encoding="utf-8"))
    modeling_contract = yaml.safe_load((FASE5_PHASE6_READY / "phase6_modeling_contract.yaml").read_text(encoding="utf-8"))
    phase6_manifest = json.loads((FASE5_PHASE6_READY / "phase6_ready_manifest.json").read_text(encoding="utf-8"))

    validate_no_split_column(phase6_feature_matrix, "phase6_feature_matrix.csv")
    validate_transform_policy(transform_params, column_groups)

    assert modeling_contract["methodology"] == "inferential_comparative_observational"
    assert modeling_contract["sample_policy"]["use_holdout_test_set"] is False
    assert modeling_contract["sample_policy"]["train_test_split_created"] is False
    assert modeling_contract["sample_policy"]["split_column_present"] is False
    assert "split_col" not in modeling_contract.get("contract", {})

    assert phase6_feature_matrix.shape == feature_matrix.shape
    assert list(phase6_feature_matrix.columns) == list(feature_matrix.columns)
    assert phase6_schema.shape[0] == feature_matrix.shape[1]
    assert phase6_missing_cols.shape[0] == feature_matrix.shape[1]
    assert phase6_missing_rows.shape[0] == feature_matrix.shape[0]
    assert len(column_groups["observed_core_40"]) == 40
    assert len(column_groups["observed_core_46"]) == 46
    assert len(column_groups["observed_core_v2_added"]) == 6
    assert len(column_groups["candidate_numeric_model_features"]) > 0
    assert modeling_contract["contract"]["no_imputation"] is True
    assert modeling_contract["contract"]["phase6_should_not_recompute_phase5"] is True
    for name, metadata in phase6_manifest["files"].items():
        path = FASE5_PHASE6_READY / metadata["path"]
        assert path.exists(), name
        assert _sha256_file(path) == metadata["sha256"], name
    checks.append("phase6_ready_contract")

    workbook = load_workbook(FASE5_OUTPUTS / "MVP_AUDITABLE.xlsx", read_only=True)
    for sheet in EXPECTED_EXCEL_SHEETS:
        assert sheet in workbook.sheetnames, f"Sheet {sheet} missing from Excel"
    assert workbook["1_Hipotesis"].max_row >= 10
    assert workbook["5_Variables_40"].max_row >= 47
    assert workbook["6_Matriz_40_Humana"].max_row == 44
    assert workbook["9_Normalizacion"].max_row >= 12
    assert workbook["11_Features_Fase6"].max_row == 44
    assert workbook["12_Diccionario_Cols"].max_row >= 100
    checks.append("excel_contract")

    assert manifest["rules"]["phase3_phase4_immutable"] is True
    assert manifest["rules"]["no_imputation"] is True
    assert manifest["phase5_implementation_hashed"]
    for name, sha in manifest["outputs"].items():
        path = FASE5_OUTPUTS / name
        if name.startswith("phase6_ready"):
            path = FASE5_PHASE6_READY / name.split("/")[1]
        assert path.exists(), name
        assert _sha256_file(path) == sha, name
    for rel_path, metadata in manifest["phase5_implementation_hashed"].items():
        path = MVP_ROOT / rel_path
        assert path.exists(), rel_path
        assert _sha256_file(path) == metadata["sha256"], rel_path
    for rel_path, metadata in manifest["inputs_hashed"].items():
        path = MVP_ROOT.parent / rel_path
        assert path.exists(), rel_path
        assert _sha256_file(path) == metadata["sha256"], rel_path
    checks.append("manifest_hashes")

    return checks


def main() -> None:
    checks = validate_phase5()
    print("Fase 5 MVP validation completed")
    for check in checks:
        print(f"- {check}: OK")


if __name__ == "__main__":
    main()
