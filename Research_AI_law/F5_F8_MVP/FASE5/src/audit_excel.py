"""Genera un Excel auditable y legible para revision humana de Fase 5."""

from __future__ import annotations

import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from _common.paths import FASE3_ROOT, FASE5_OUTPUTS
from .sample import get_mvp_entities_detail
from .variables import build_variable_catalog, get_mvp_variables


META_COLS_PUBLIC = ["iso3", "country_name_canonical", "region", "income_group"]

ROLE_COLORS = {
    "X1_regulatory": "F4CCCC",
    "Y_Q1_investment": "CFE2F3",
    "Y_Q1_or_control": "E2F0D9",
    "Y_Q2_adoption": "D9EAD3",
    "Y_Q3_innovation": "D9D2E9",
    "Y_Q5_population_usage": "DDEBF7",
    "Y_Q6_public_sector": "E4DFEC",
    "Y_Q6_public_sector_aux": "D9E2F3",
    "X2_control": "FFF2CC",
}

COLUMN_TYPE_COLORS = {
    "metadato": "B7DEE8",
    "variable_observada_46": "FFF2CC",
    "derivada_log": "D9EAD3",
    "derivada_zscore": "FCE5CD",
    "agregado_regulatorio": "EADCF8",
    "one_hot_categorica": "F4CCCC",
    "auditoria": "EFEFEF",
}

INCOME_GROUP_DEFINITIONS = {
    "High income": "Pais clasificado como ingreso alto en la metadata Fase 3 / World Bank.",
    "Upper middle income": "Pais clasificado como ingreso medio alto en la metadata Fase 3 / World Bank.",
    "Lower middle income": "Pais clasificado como ingreso medio bajo en la metadata Fase 3 / World Bank.",
    "Low income": "Pais clasificado como ingreso bajo en la metadata Fase 3 / World Bank.",
}


def _safe_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value)


def _role_label(role: str) -> str:
    labels = {
        "X1_regulatory": "Tratamiento regulatorio IA",
        "Y_Q1_investment": "Pregunta 1: inversion / ecosistema",
        "Y_Q1_or_control": "Pregunta 1 o control macro",
        "Y_Q2_adoption": "Pregunta 2: adopcion / difusion",
        "Y_Q3_innovation": "Pregunta 3: innovacion / capacidad",
        "Y_Q5_population_usage": "Pregunta 5: uso IA poblacion",
        "Y_Q6_public_sector": "Pregunta 6: uso IA sector publico",
        "Y_Q6_public_sector_aux": "Pregunta 6: auxiliar sector publico",
        "X2_control": "Controles socioeconomicos, institucionales y tecnologicos",
    }
    return labels.get(role, role)


def _role_plain_explanation(role: str) -> str:
    labels = {
        "X1_regulatory": "Variables que describen el tipo o intensidad de regulacion de IA.",
        "Y_Q1_investment": "Variables que aproximan inversion privada y ecosistema empresarial IA.",
        "Y_Q1_or_control": "Variable que puede informar inversion macro o usarse como control.",
        "Y_Q2_adoption": "Variables que aproximan adopcion y difusion de IA.",
        "Y_Q3_innovation": "Variables que aproximan innovacion, investigacion o capacidad IA.",
        "Y_Q5_population_usage": "Variables que aproximan uso de IA por la poblacion.",
        "Y_Q6_public_sector": "Variables que aproximan uso o capacidad IA en sector publico.",
        "Y_Q6_public_sector_aux": "Variable auxiliar de sector publico con menor cobertura.",
        "X2_control": "Controles para no confundir regulacion con riqueza, instituciones o infraestructura.",
    }
    return labels.get(role, "")


def _variable_short_name(variable: str, reason: str) -> str:
    manual = {
        "iapp_ley_ia_vigente": "Ley IA vigente",
        "iapp_categoria_obligatoriedad": "Tipo de obligatoriedad",
        "iapp_proyecto_ley_ia": "Proyecto de ley IA",
        "iapp_modelo_gobernanza": "Modelo de gobernanza",
        "iapp_n_leyes_relacionadas": "N leyes relacionadas",
        "iapp_n_autoridades": "N autoridades IA",
        "oxford_ind_company_investment_emerging_tech": "Inversion empresarial emergente",
        "oxford_ind_ai_unicorns_log": "Unicornios IA",
        "oxford_ind_non_ai_unicorns_log": "Unicornios no IA",
        "oxford_ind_vc_availability": "Disponibilidad VC",
        "wipo_c_vencapdeal_score": "Deals capital riesgo",
        "wb_fdi_net_inflows": "FDI neto",
        "ms_h2_2025_ai_diffusion_pct": "Difusion IA MS H2 2025",
        "ms_h1_2025_ai_diffusion_pct": "Difusion IA MS H1 2025",
        "ms_change_pp": "Cambio difusion MS",
        "anthropic_usage_pct": "Uso IA Anthropic",
        "anthropic_collaboration_pct": "Colaboracion Anthropic",
        "oecd_5_ict_business_oecd_biz_ai_pct": "Empresas usan IA OECD",
        "oecd_5_ict_business_oecd_biz_bigdata_pct": "Empresas usan Big Data OECD",
        "oxford_public_sector_adoption": "Adopcion sector publico",
        "oxford_ind_adoption_emerging_tech": "Adopcion tech emergente",
        "wipo_out_score": "Output innovacion WIPO",
        "oxford_total_score": "Score total Oxford",
        "wipo_gii_score": "Global Innovation Index",
        "oxford_innovation_capacity": "Capacidad innovacion Oxford",
        "oxford_ind_ai_research_papers_log": "Papers IA Oxford",
        "stanford_fig_6_3_5_fig_6_3_5_volume_of_publications": "Publicaciones IA Stanford",
        "wb_patent_applications_residents": "Patentes residentes",
        "wb_gdp_per_capita_ppp": "PIB per capita PPP",
        "wb_internet_penetration": "Penetracion internet",
        "wb_tertiary_education_enrollment": "Educacion terciaria",
        "wb_rd_expenditure_pct_gdp": "Gasto I+D PIB",
        "wb_researchers_rd_per_million": "Investigadores I+D",
        "wb_government_effectiveness": "Efectividad gobierno",
        "wb_rule_of_law": "Estado de derecho",
        "wb_regulatory_quality": "Calidad regulatoria",
        "wipo_in_score": "Inputs innovacion WIPO",
        "wb_mobile_subscriptions_per100": "Suscripciones moviles",
        "wb_secure_servers_per_1m": "Servidores seguros",
        "wb_electric_consumption_kwh_pc": "Consumo electrico",
        "oxford_e_government_delivery": "Entrega gobierno electronico",
        "oxford_government_digital_policy": "Politica digital gobierno",
        "oxford_ind_data_governance": "Gobernanza datos",
        "oxford_governance_ethics": "Etica gobernanza IA",
        "oecd_2_indigo_oecd_indigo_score": "OECD INDIGO",
        "oecd_4_digital_gov_oecd_digital_gov_overall": "OECD Digital Gov",
    }
    return manual.get(variable, str(reason).split(";")[0][:45])


def _transform_explanation(transform: str) -> str:
    return {
        "none": "Sin transformacion tecnica: la variable observada se usa tal como viene.",
        "one_hot": "Variable categorica: se conserva el texto y, para modelos, se crean columnas 0/1 por categoria.",
        "zscore": "Ademas del valor original, Fase 5 crea una version estandarizada robusta para comparar escalas.",
        "log_zscore": "Ademas del valor original, Fase 5 crea log y z-score robusto para reducir escala/asimetria.",
        "signed_log_zscore": "Similar a log_zscore, pero permite valores negativos conservando el signo.",
    }.get(transform, "Transformacion configurada en Fase 5.")


def _with_variable_ids(catalog: pd.DataFrame) -> pd.DataFrame:
    out = catalog.copy()
    out.insert(0, "id_variable", [f"V{i:02d}" for i in range(1, len(out) + 1)])
    out.insert(2, "nombre_corto_para_excel", [
        _variable_short_name(row.variable_matriz, row.razon)
        for row in out.itertuples(index=False)
    ])
    return out


def build_countries_sheet(wide_mvp: pd.DataFrame) -> pd.DataFrame:
    detail = get_mvp_entities_detail()
    cols = ["iso3", "country_name_canonical", "region", "income_group", "pct_variables_available"]
    countries = wide_mvp[[c for c in cols if c in wide_mvp.columns]].copy()
    out = countries.merge(detail, on="iso3", how="left")
    out["income_group_explicacion"] = out["income_group"].map(INCOME_GROUP_DEFINITIONS).fillna(
        "Sin clasificacion de ingreso disponible en la metadata de Fase 3."
    )
    out["nota_para_auditor"] = ""
    out.loc[out["iso3"].eq("TWN"), "nota_para_auditor"] = (
        "Taiwan se mantiene por relevancia IA y ranking Microsoft Top 30, "
        "pero Fase 3 no trae region ni income_group comparable para TWN. "
        "Los vacios son faltantes reales, no error ni imputacion."
    )
    out = out.rename(columns={
        "iso3": "codigo_iso3",
        "country_name_canonical": "pais",
        "region": "region_geografica",
        "income_group": "grupo_ingreso",
        "pct_variables_available": "pct_variables_disponibles_matriz_madre",
        "category": "categoria_muestra_mvp",
        "rank_ms": "ranking_microsoft_si_aplica",
        "reason": "motivo_inclusion_mvp",
    })
    return out[
        [
            "codigo_iso3",
            "pais",
            "region_geografica",
            "grupo_ingreso",
            "income_group_explicacion",
            "categoria_muestra_mvp",
            "ranking_microsoft_si_aplica",
            "motivo_inclusion_mvp",
            "pct_variables_disponibles_matriz_madre",
            "nota_para_auditor",
        ]
    ]


def build_variables_sheet() -> pd.DataFrame:
    catalog = _with_variable_ids(build_variable_catalog())
    out = catalog.copy()
    out["rol_lectura_humana"] = out["rol_mvp"].map(_role_label)
    out["explicacion_del_rol"] = out["rol_mvp"].map(_role_plain_explanation)
    out["incluida_en_matriz_46_observadas"] = "SI"
    out["que_significa_transform"] = out["transform"].map(_transform_explanation)
    out["donde_auditar_en_excel"] = (
        "Hoja 6_Matriz_40_Humana por id_variable; hoja 13_Trazabilidad filtra por variable_codigo_original."
    )
    keep = [
        "id_variable",
        "nombre_corto_para_excel",
        "variable_matriz",
        "incluida_en_matriz_46_observadas",
        "rol_lectura_humana",
        "explicacion_del_rol",
        "subpregunta",
        "tipo",
        "unit",
        "direction",
        "transform",
        "added_in_version",
        "tier",
        "que_significa_transform",
        "razon",
        "donde_auditar_en_excel",
        "source_id",
        "table_id",
        "pct_complete",
        "n_countries_available",
        "known_limitations",
    ]
    return out[[c for c in keep if c in out.columns]]


def build_observed_matrix_sheet(feature_matrix: pd.DataFrame) -> pd.DataFrame:
    variables = build_variables_sheet()
    label_by_var = dict(zip(
        variables["variable_matriz"],
        variables["id_variable"] + " - " + variables["nombre_corto_para_excel"],
    ))
    cols = [c for c in META_COLS_PUBLIC if c in feature_matrix.columns] + get_mvp_variables()
    out = feature_matrix[cols].copy()
    renamed = {
        "iso3": "codigo_iso3",
        "country_name_canonical": "pais",
        "region": "region_geografica",
        "income_group": "grupo_ingreso",
    }
    renamed.update(label_by_var)
    return out.rename(columns=renamed)


def build_v2_variables_detail_sheet() -> pd.DataFrame:
    variables = build_variables_sheet()
    if "added_in_version" not in variables.columns:
        return pd.DataFrame()
    added = variables[variables["added_in_version"].astype(str).eq("2.0")].copy()
    if added.empty:
        return pd.DataFrame()
    coverage = {
        "oxford_e_government_delivery": "43/43",
        "oxford_government_digital_policy": "43/43",
        "oxford_ind_data_governance": "43/43",
        "oxford_governance_ethics": "43/43",
        "oecd_2_indigo_oecd_indigo_score": "43/43",
        "oecd_4_digital_gov_oecd_digital_gov_overall": "31/43",
    }
    added.insert(0, "numero_v2", range(41, 41 + len(added)))
    added["cobertura_mvp_v1"] = added["variable_matriz"].map(coverage)
    added["nota_v2"] = added["variable_matriz"].map({
        "oecd_4_digital_gov_oecd_digital_gov_overall": "Auxiliary tier por cobertura 31/43.",
    }).fillna("Primary Q6 support.")
    return added


def build_traceability_sheet() -> pd.DataFrame:
    trace_path = FASE3_ROOT / "outputs" / "matriz_madre_trazabilidad.csv"
    trace = pd.read_csv(trace_path)
    iso3 = set(get_mvp_entities_detail()["iso3"])
    vars_ = set(get_mvp_variables())
    keep_cols = [
        "cell_id", "iso3", "country_name_canonical", "variable_matriz",
        "source_id", "table_id", "original_variable", "year", "value_original",
        "value_numeric", "value_text", "unit", "confidence_level",
        "source_file", "source_sheet", "row_identifier",
    ]
    filtered = trace[trace["iso3"].isin(iso3) & trace["variable_matriz"].isin(vars_)].copy()
    out = filtered[[c for c in keep_cols if c in filtered.columns]]
    return out.rename(columns={
        "iso3": "codigo_iso3",
        "country_name_canonical": "pais",
        "variable_matriz": "variable_fase5",
        "source_id": "fuente",
        "table_id": "tabla_fuente",
        "original_variable": "nombre_original_en_fuente",
        "year": "anio",
        "value_original": "valor_original",
        "value_numeric": "valor_numerico",
        "value_text": "valor_texto",
        "unit": "unidad",
        "confidence_level": "nivel_confianza",
        "source_file": "archivo_fuente",
        "source_sheet": "hoja_fuente",
        "row_identifier": "fila_origen",
    })


def _source_for_zscore(col: str, feature_columns: set[str]) -> str:
    source = col[:-2]
    if source in feature_columns:
        return source
    return ""


def _source_for_log(col: str, observed_vars: set[str]) -> str:
    for var in sorted(observed_vars, key=len, reverse=True):
        if col == f"{var}_log":
            return var
    return ""


def classify_column(col: str, catalog_by_var: dict[str, dict], feature_columns: set[str]) -> dict:
    observed_vars = set(catalog_by_var)
    if col in observed_vars:
        row = catalog_by_var[col]
        return {
            "tipo_columna": "variable_observada_46",
            "origen": "Variable observada real de Fase 3, seleccionada en Fase 5.",
            "variable_fuente": col,
            "rol_mvp": row.get("rol_mvp", ""),
            "lectura_humana": row.get("razon", ""),
            "parte_estudio_real_43x46": "SI",
        }
    if col in META_COLS_PUBLIC or col in {
        "entity_type", "n_sources_present", "source_list", "n_variables_available",
        "pct_variables_available", "included_in_matrix", "included_in_dense_candidate",
        "inclusion_notes",
    }:
        explanations = {
            "iso3": "Codigo internacional ISO3 del pais.",
            "country_name_canonical": "Nombre canonico del pais.",
            "region": "Region geografica proveniente de la metadata Fase 3.",
            "income_group": "Grupo de ingreso pais, segun metadata Fase 3 / World Bank.",
            "entity_type": "Tipo de entidad; en Fase 5 todos son paises ISO3.",
            "n_sources_present": "Numero de fuentes con datos disponibles para el pais.",
            "source_list": "Lista de fuentes presentes para el pais.",
            "n_variables_available": "Cantidad de variables disponibles en la Matriz Madre.",
            "pct_variables_available": "Porcentaje de variables disponibles en la Matriz Madre.",
            "included_in_matrix": "Indicador de inclusion en Matriz Madre.",
            "included_in_dense_candidate": "Indicador auxiliar de densidad usado en Fases previas.",
            "inclusion_notes": "Notas de inclusion heredadas de Fase 3 si existen.",
        }
        return {
            "tipo_columna": "metadato",
            "origen": "Metadata de pais / auditoria; no cuenta dentro de las 46 variables observadas.",
            "variable_fuente": "",
            "rol_mvp": "",
            "lectura_humana": explanations.get(col, "Metadato auxiliar."),
            "parte_estudio_real_43x46": "NO",
        }
    log_source = _source_for_log(col, observed_vars)
    if log_source:
        return {
            "tipo_columna": "derivada_log",
            "origen": "Transformacion logaritmica creada en Fase 5 para modelado.",
            "variable_fuente": log_source,
            "rol_mvp": catalog_by_var[log_source].get("rol_mvp", ""),
            "lectura_humana": "Reduce asimetria/escala de una variable numerica; conserva NaN.",
            "parte_estudio_real_43x46": "NO, derivada de una variable observada.",
        }
    if col.endswith("_z"):
        source = _source_for_zscore(col, feature_columns)
        role = catalog_by_var.get(source, {}).get("rol_mvp", "")
        if not role:
            log_parent = _source_for_log(source, observed_vars)
            role = catalog_by_var.get(log_parent, {}).get("rol_mvp", "")
        return {
            "tipo_columna": "derivada_zscore",
            "origen": "Estandarizacion robusta creada en Fase 5 para comparar escalas.",
            "variable_fuente": source,
            "rol_mvp": role,
            "lectura_humana": "Valor centrado por mediana y escalado por MAD; conserva NaN.",
            "parte_estudio_real_43x46": "NO, derivada de una variable observada.",
        }
    if col in {
        "n_binding", "n_non_binding", "n_hybrid", "regulatory_intensity",
        "n_regulatory_mechanisms",
    }:
        return {
            "tipo_columna": "agregado_regulatorio",
            "origen": "Feature agregado creado con taxonomia revisada en Fase 4.",
            "variable_fuente": "binding_taxonomy.yaml",
            "rol_mvp": "X1_regulatory",
            "lectura_humana": "Resume presencia/intensidad de mecanismos regulatorios para Fase 6.",
            "parte_estudio_real_43x46": "NO, feature derivada para modelado.",
        }
    if col.endswith("_variables_used"):
        return {
            "tipo_columna": "auditoria",
            "origen": "Lista de variables usadas para calcular agregados regulatorios.",
            "variable_fuente": "binding_taxonomy.yaml",
            "rol_mvp": "X1_regulatory",
            "lectura_humana": "Columna de trazabilidad; no es predictor numerico principal.",
            "parte_estudio_real_43x46": "NO",
        }
    if col.startswith("iapp_categoria_obligatoriedad_") or col.startswith("iapp_modelo_gobernanza_"):
        return {
            "tipo_columna": "one_hot_categorica",
            "origen": "Codificacion binaria de variable categorica regulatoria.",
            "variable_fuente": re.sub(r"_(binding|nonbinding|draft|strategy).*$", "", col),
            "rol_mvp": "X1_regulatory",
            "lectura_humana": "Vale 1 si el pais pertenece a esa categoria y 0 si no.",
            "parte_estudio_real_43x46": "NO, codificacion tecnica para Fase 6.",
        }
    return {
        "tipo_columna": "auditoria",
        "origen": "Columna auxiliar.",
        "variable_fuente": "",
        "rol_mvp": "",
        "lectura_humana": "Revisar segun contexto.",
        "parte_estudio_real_43x46": "NO",
    }


def build_column_dictionary(feature_matrix: pd.DataFrame) -> pd.DataFrame:
    catalog = build_variable_catalog()
    catalog_by_var = catalog.set_index("variable_matriz").to_dict("index")
    feature_columns = set(feature_matrix.columns)
    rows = []
    for i, col in enumerate(feature_matrix.columns, start=1):
        classified = classify_column(col, catalog_by_var, feature_columns)
        rows.append({
            "orden_en_feature_matrix": i,
            "columna": col,
            **classified,
        })
    return pd.DataFrame(rows)


def build_hypothesis_sheet() -> pd.DataFrame:
    rows = [
        {
            "bloque": "Hipotesis y preguntas",
            "item": "Hipotesis principal",
            "explicacion_didactica": (
                "Existe una asociacion estadisticamente significativa entre las caracteristicas "
                "de la regulacion de IA de un pais y el desarrollo de su ecosistema de IA, "
                "despues de controlar por factores socioeconomicos e institucionales?"
            ),
            "rol_en_el_proyecto": (
                "Fase 5 no prueba la hipotesis. Prepara la muestra, las variables regulatorias, "
                "los outcomes y los controles para que Fase 6 la modele."
            ),
            "fase_o_hoja_relacionada": "5_Variables_40 y 6_Matriz_40_Humana",
        },
        {
            "bloque": "Hipotesis y preguntas",
            "item": "Q1 - Inversion",
            "explicacion_didactica": "Mas regulacion restrictiva se asocia con menor inversion privada en IA?",
            "rol_en_el_proyecto": "Selecciona variables de inversion, capital de riesgo, unicornios y FDI.",
            "fase_o_hoja_relacionada": "Variables con rol Pregunta 1 en 5_Variables_40.",
        },
        {
            "bloque": "Hipotesis y preguntas",
            "item": "Q2 - Adopcion",
            "explicacion_didactica": "Que enfoque regulatorio se asocia con mayor adopcion de IA?",
            "rol_en_el_proyecto": "Selecciona variables Microsoft, Anthropic, OECD y Oxford sobre adopcion/difusion.",
            "fase_o_hoja_relacionada": "Variables con rol Pregunta 2 en 5_Variables_40.",
        },
        {
            "bloque": "Hipotesis y preguntas",
            "item": "Q3 - Innovacion",
            "explicacion_didactica": "Hay relacion entre regulacion de IA e indicadores de innovacion?",
            "rol_en_el_proyecto": "Selecciona variables WIPO, Oxford, Stanford, papers y patentes.",
            "fase_o_hoja_relacionada": "Variables con rol Pregunta 3 en 5_Variables_40.",
        },
        {
            "bloque": "Hipotesis y preguntas",
            "item": "Q4 - Perfil regulatorio",
            "explicacion_didactica": "Que tipos regulatorios emergen al agrupar paises por perfil legal?",
            "rol_en_el_proyecto": (
                "MVP no usa NLP legal. Usa variables estructuradas IAPP y agregados de la taxonomia Fase 4."
            ),
            "fase_o_hoja_relacionada": "Variables regulatorias y columnas tecnicas en 11_Features_Fase6.",
        },
        {
            "bloque": "Hipotesis y preguntas",
            "item": "Q5 - Uso IA poblacion",
            "explicacion_didactica": "Como se asocia la regulacion de IA con uso de IA por la poblacion?",
            "rol_en_el_proyecto": "Reusa variables Anthropic y Oxford existentes para modelado separado en Fase 6.",
            "fase_o_hoja_relacionada": "Contrato tecnico phase6_modeling_contract.yaml.",
        },
        {
            "bloque": "Hipotesis y preguntas",
            "item": "Q6 - Uso IA sector publico",
            "explicacion_didactica": "Como se asocia la regulacion de IA con uso/capacidad de IA en sector publico?",
            "rol_en_el_proyecto": "Agrega 6 variables Oxford/OECD para consistencia metodologica de Fase 6.",
            "fase_o_hoja_relacionada": "5b_Variables_46_Detalle y Muestra_Analitica_v2_1.",
        },
    ]
    rows.extend([
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 1 - Entendimiento del problema y EDA preliminar raw",
            "explicacion_didactica": (
                "Traduce la pregunta legislativa sobre IA en preguntas analizables con datos. "
                "Identifica stakeholders, riesgos, supuestos y fuentes iniciales."
            ),
            "rol_en_el_proyecto": "Equivale al inicio de CRISP-DM: Business/Policy Understanding y Data Understanding.",
            "fase_o_hoja_relacionada": "CONTEXTOS/1.PLAN_EDA_PRELIMINAR_RAW.md",
        },
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 2 - Transicion y diseno de Matriz Madre",
            "explicacion_didactica": (
                "Ordena las fuentes, define entidades, variables, trazabilidad y criterios para pasar "
                "de archivos dispersos a una estructura auditable."
            ),
            "rol_en_el_proyecto": "Puente entre Data Understanding y Data Preparation.",
            "fase_o_hoja_relacionada": "CONTEXTOS/2.TRANSICION_FASE2_A_FASE3_MATRIZ_MADRE.md",
        },
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 3 - Matriz Madre",
            "explicacion_didactica": (
                "Construye la tabla pais x atributo con datos trazables, diccionario de variables "
                "y control de calidad de celdas."
            ),
            "rol_en_el_proyecto": "Data Preparation base: limpieza, estandarizacion inicial y trazabilidad.",
            "fase_o_hoja_relacionada": "FASE3/outputs",
        },
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 4 - EDA principal",
            "explicacion_didactica": (
                "Explora cobertura, distribuciones, outliers, redundancias, submuestras y viabilidad "
                "de las preguntas antes de modelar."
            ),
            "rol_en_el_proyecto": "Data Understanding avanzado y control metodologico antes de modelado.",
            "fase_o_hoja_relacionada": "FASE4/outputs/eda_principal",
        },
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 5 - Preparacion de datos MVP",
            "explicacion_didactica": (
                "Define 43 paises, selecciona 46 variables reales, crea normalizaciones/encoding, "
                "documenta faltantes y entrega la matriz tecnica para Fase 6."
            ),
            "rol_en_el_proyecto": "CRISP-DM Data Preparation: feature vector listo para algoritmos.",
            "fase_o_hoja_relacionada": "Este Excel y outputs/phase6_ready",
        },
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 6 - Modelado",
            "explicacion_didactica": (
                "Aplica regresion, clasificacion, modelos penalizados o enfoques causales exploratorios "
                "para estimar asociaciones entre regulacion y ecosistema IA."
            ),
            "rol_en_el_proyecto": "CRISP-DM Modeling.",
            "fase_o_hoja_relacionada": "Se alimenta desde outputs/phase6_ready",
        },
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 7 - Evaluacion, sensibilidad y robustez",
            "explicacion_didactica": (
                "Evalua si los resultados son estables, interpretables y mejores que baselines; "
                "controla sobreajuste, falsos positivos y sensibilidad a faltantes."
            ),
            "rol_en_el_proyecto": "CRISP-DM Evaluation.",
            "fase_o_hoja_relacionada": "Fase posterior a modelado.",
        },
        {
            "bloque": "Mapa de 8 fases del proyecto",
            "item": "Fase 8 - Reporte ejecutivo y despliegue",
            "explicacion_didactica": (
                "Convierte resultados tecnicos en informe, narrativa de politica publica, anexos "
                "auditables y recomendaciones para decision legislativa."
            ),
            "rol_en_el_proyecto": "CRISP-DM Deployment adaptado a politica publica.",
            "fase_o_hoja_relacionada": "Reporte ejecutivo MVP final.",
        },
    ])
    return pd.DataFrame(rows)


def build_audit_protocol_sheet() -> pd.DataFrame:
    n_vars = len(get_mvp_variables())
    rows = [
        {
            "paso": 1,
            "accion_humana": "Leer 0_Leer_Primero y 1_Hipotesis.",
            "que_debe_quedar_claro": "La Fase 5 prepara datos; no entrega aun resultados causales ni modelos.",
            "hoja_excel": "0_Leer_Primero / 1_Hipotesis",
        },
        {
            "paso": 2,
            "accion_humana": "Revisar que los 43 paises tengan sentido para el MVP.",
            "que_debe_quedar_claro": "Chile, peers LATAM, lideres IA y comparadores relevantes estan incluidos.",
            "hoja_excel": "3_Paises_43",
        },
        {
            "paso": 3,
            "accion_humana": f"Elegir una variable del estudio real usando su ID V01-V{n_vars:02d}.",
            "que_debe_quedar_claro": "Cada variable tiene nombre corto, codigo exacto, rol, fuente y explicacion.",
            "hoja_excel": "5_Variables_40",
        },
        {
            "paso": 4,
            "accion_humana": "Buscar esa variable por ID en la matriz humana.",
            "que_debe_quedar_claro": f"La hoja 6 contiene los 43 paises y las {n_vars} variables observadas; no incluye features tecnicas.",
            "hoja_excel": "6_Matriz_40_Humana",
        },
        {
            "paso": 5,
            "accion_humana": "Auditar el origen de una celda concreta.",
            "que_debe_quedar_claro": "Filtrar por pais y por variable_codigo_original para ver fuente, tabla y valor original.",
            "hoja_excel": "13_Trazabilidad",
        },
        {
            "paso": 6,
            "accion_humana": "Revisar vacios y paises con metadata incompleta.",
            "que_debe_quedar_claro": "Los vacios no fueron imputados. Taiwan tiene vacios esperados por cobertura/metadata.",
            "hoja_excel": "8_Casos_Atencion",
        },
        {
            "paso": 7,
            "accion_humana": "Dejar la matriz tecnica para el equipo de modelado.",
            "que_debe_quedar_claro": "La hoja 11 tiene columnas porque agrega transformaciones para Fase 6. Sin embargo, no hay divisiones predictivas artificiales.",
            "hoja_excel": "11_Features_Fase6",
        },
    ]
    return pd.DataFrame(rows)


def build_income_region_sheet(feature_matrix: pd.DataFrame) -> pd.DataFrame:
    countries = build_countries_sheet(feature_matrix)
    group_rows = []
    for group, definition in INCOME_GROUP_DEFINITIONS.items():
        n = int(countries["grupo_ingreso"].eq(group).sum())
        group_rows.append({
            "tipo": "income_group",
            "categoria": group,
            "n_paises_mvp": n,
            "explicacion": definition,
        })
    missing_income = int(countries["grupo_ingreso"].isna().sum() + countries["grupo_ingreso"].eq("").sum())
    group_rows.append({
        "tipo": "income_group",
        "categoria": "Sin clasificacion disponible",
        "n_paises_mvp": missing_income,
        "explicacion": (
            "No hay grupo de ingreso comparable en la metadata Fase 3. "
            "En esta muestra corresponde a Taiwan."
        ),
    })
    for region, n in countries["region_geografica"].fillna("Sin region disponible").value_counts().items():
        group_rows.append({
            "tipo": "region",
            "categoria": region,
            "n_paises_mvp": int(n),
            "explicacion": "Region geografica heredada de la metadata Fase 3.",
        })
    return pd.DataFrame(group_rows)


def build_color_legend_sheet() -> pd.DataFrame:
    rows = [
        {
            "color_hex": ROLE_COLORS["X1_regulatory"],
            "aparece_en": "6_Matriz_40_Humana",
            "significado": "Tratamiento regulatorio IA",
            "detalle": "Variables IAPP que describen ley IA, obligatoriedad, gobernanza e institucionalidad.",
        },
        {
            "color_hex": ROLE_COLORS["Y_Q1_investment"],
            "aparece_en": "6_Matriz_40_Humana",
            "significado": "Q1 inversion / ecosistema",
            "detalle": "Outcomes o proxies de inversion privada, VC, unicornios y capital.",
        },
        {
            "color_hex": ROLE_COLORS["Y_Q2_adoption"],
            "aparece_en": "6_Matriz_40_Humana",
            "significado": "Q2 adopcion / difusion",
            "detalle": "Uso de IA en empresas, sector publico y difusion observada.",
        },
        {
            "color_hex": ROLE_COLORS["Y_Q3_innovation"],
            "aparece_en": "6_Matriz_40_Humana",
            "significado": "Q3 innovacion",
            "detalle": "Innovacion, papers, patentes, scores Oxford/WIPO/Stanford.",
        },
        {
            "color_hex": ROLE_COLORS["X2_control"],
            "aparece_en": "6_Matriz_40_Humana",
            "significado": "Controles",
            "detalle": "PIB, internet, educacion, I+D, instituciones, infraestructura.",
        },
        {
            "color_hex": COLUMN_TYPE_COLORS["metadato"],
            "aparece_en": "6_Matriz_40_Humana y 11_Features_Fase6",
            "significado": "Metadatos",
            "detalle": "Identificadores de pais; no cuentan dentro de las 46 variables.",
        },
        {
            "color_hex": COLUMN_TYPE_COLORS["derivada_log"],
            "aparece_en": "11_Features_Fase6",
            "significado": "Transformacion log",
            "detalle": "Columna tecnica derivada para modelado; no es variable nueva observada.",
        },
        {
            "color_hex": COLUMN_TYPE_COLORS["derivada_zscore"],
            "aparece_en": "11_Features_Fase6",
            "significado": "Z-score robusto",
            "detalle": "Columna tecnica escalada por mediana/MAD para comparar variables.",
        },
        {
            "color_hex": COLUMN_TYPE_COLORS["agregado_regulatorio"],
            "aparece_en": "11_Features_Fase6",
            "significado": "Agregado regulatorio",
            "detalle": "Feature creado con taxonomia Fase 4 binding/non-binding/hybrid.",
        },
    ]
    return pd.DataFrame(rows)


def build_data_attention_sheet(feature_matrix: pd.DataFrame) -> pd.DataFrame:
    observed = get_mvp_variables()
    n_observed = len(observed)
    rows = []
    country_cols = ["iso3", "country_name_canonical", "region", "income_group"]
    for _, row in feature_matrix.iterrows():
        missing_40 = int(row[observed].isna().sum())
        if missing_40 >= 10 or pd.isna(row.get("region")) or pd.isna(row.get("income_group")):
            note = "Vacios relevantes en variables observadas."
            if row["iso3"] == "TWN":
                note = (
                    "Taiwan: incluido por relevancia IA y ranking Microsoft Top 30 #23. "
                    "No tiene region ni income_group comparable en Fase 3; sus vacios son falta de dato fuente, no error."
                )
            rows.append({
                "codigo_iso3": row["iso3"],
                "pais": row["country_name_canonical"],
                "region": row.get("region"),
                "income_group": row.get("income_group"),
                "variables_observadas_vacias_de_46": missing_40,
                "variables_observadas_disponibles_de_46": n_observed - missing_40,
                "pct_variables_disponibles_matriz_madre": row.get("pct_variables_available"),
                "nota_para_revision": note,
            })
    return pd.DataFrame(rows).sort_values(
        ["variables_observadas_vacias_de_46", "codigo_iso3"], ascending=[False, True]
    )


def build_normalization_sheet(transform_params: pd.DataFrame | None) -> pd.DataFrame:
    if transform_params is None or transform_params.empty:
        n_log = 0
        n_z = 0
    else:
        n_log = int(transform_params["transform_type"].eq("log_transform").sum())
        n_z = int(transform_params["transform_type"].eq("robust_zscore").sum())
    rows = [
        {
            "tema": "Por que normalizar",
            "explicacion_didactica": (
                "El estudio mezcla unidades muy distintas: dolares, porcentajes, scores 0-100, "
                "conteos de papers/patentes, variables binarias y categorias regulatorias. "
                "Si se modelan sin preparar, una variable con numeros grandes puede dominar "
                "artificialmente a otra igual de importante pero medida en porcentaje."
            ),
            "decision_fase5": "Crear versiones tecnicas comparables sin borrar las variables originales.",
            "donde_ver_detalle": "11_Features_Fase6 y 14_Transformaciones",
        },
        {
            "tema": "Que significa nivel universal",
            "explicacion_didactica": (
                "No significa convertir todos los paises a una misma realidad economica. "
                "Significa poner variables numericas en escalas comparables para que el modelo "
                "no confunda magnitud de unidad con importancia sustantiva."
            ),
            "decision_fase5": "Usar transformaciones reproducibles y documentadas.",
            "donde_ver_detalle": "10_Cobertura / 14_Transformaciones",
        },
        {
            "tema": "Variables originales preservadas",
            "explicacion_didactica": (
                "Fase 5 no reemplaza la evidencia original. La matriz humana mantiene las 46 "
                "variables observadas tal como vienen de Fase 3; la matriz tecnica agrega columnas "
                "derivadas para modelado."
            ),
            "decision_fase5": "Preservar originales y agregar derivadas.",
            "donde_ver_detalle": "6_Matriz_40_Humana vs 11_Features_Fase6",
        },
        {
            "tema": "Transformacion logaritmica",
            "explicacion_didactica": (
                "Se usa cuando una variable tiene distribucion muy asimetrica: pocos paises concentran "
                "valores enormes y muchos paises tienen valores pequenos. El log reduce esa asimetria "
                "para que potencias como EE.UU. o China no dominen mecanicamente."
            ),
            "decision_fase5": f"Se generaron {n_log} transformaciones log/signed_log.",
            "donde_ver_detalle": "14_Transformaciones",
        },
        {
            "tema": "signed_log1p",
            "explicacion_didactica": (
                "Algunas variables pueden ser negativas, como flujos netos de inversion extranjera. "
                "En esos casos no se puede aplicar log comun directamente. signed_log1p reduce escala "
                "sin perder el signo positivo/negativo."
            ),
            "decision_fase5": "Aplicar signed_log1p cuando hay valores negativos.",
            "donde_ver_detalle": "14_Transformaciones, columna method",
        },
        {
            "tema": "Z-score robusto",
            "explicacion_didactica": (
                "Convierte variables numericas a una escala comparable usando mediana y MAD "
                "(desviacion absoluta mediana). Es mas resistente a outliers que el z-score clasico "
                "basado en media y desviacion estandar."
            ),
            "decision_fase5": f"Se generaron {n_z} columnas robust_zscore.",
            "donde_ver_detalle": "14_Transformaciones",
        },
        {
            "tema": "Formula del z-score robusto",
            "explicacion_didactica": "z_robusto = (valor - mediana_de_la_variable) / MAD_de_la_variable.",
            "decision_fase5": "Guardar mediana y MAD para auditoria/reproducibilidad.",
            "donde_ver_detalle": "14_Transformaciones, columnas median y mad",
        },
        {
            "tema": "Variables categoricas",
            "explicacion_didactica": (
                "Variables como tipo de obligatoriedad o modelo de gobernanza son texto/categoria. "
                "Los modelos necesitan numeros, por eso se codifican en columnas 0/1 por categoria."
            ),
            "decision_fase5": "Aplicar one-hot encoding y preservar categoria original.",
            "donde_ver_detalle": "5_Variables_40, 11_Features_Fase6, 12_Diccionario_Cols",
        },
        {
            "tema": "Que NO corrige la normalizacion",
            "explicacion_didactica": (
                "Normalizar no elimina sesgos de fuente, no prueba causalidad, no inventa datos faltantes "
                "y no garantiza ausencia de falsos positivos. Solo reduce problemas de escala y asimetria."
            ),
            "decision_fase5": "Registrar esta limitacion explicitamente.",
            "donde_ver_detalle": "1_Hipotesis y fases 6-7",
        },
        {
            "tema": "Datos vacios",
            "explicacion_didactica": (
                "Un vacio no es cero. Puede significar que el pais no reporta, que la fuente no cubre "
                "ese territorio, o que el dato no es comparable. Rellenarlo sin justificacion puede "
                "crear evidencia artificial."
            ),
            "decision_fase5": "No imputar en Fase 5; preservar NaN.",
            "donde_ver_detalle": "8_Casos_Atencion, 10_Cobertura, phase6_missingness_*",
        },
        {
            "tema": "Que queda para Fase 6",
            "explicacion_didactica": (
                "La validacion del estudio se realizara posteriormente mediante validacion interna, bootstrap, cross-validation diagnostica y analisis de sensibilidad en Fase 6/Fase 7. Cualquier score por pais en fases posteriores debe interpretarse como posicionamiento descriptivo in-sample, no como prediccion independiente."
            ),
            "decision_fase5": "Entregar matriz y missingness; no forzar una estrategia unica.",
            "donde_ver_detalle": "outputs/phase6_ready",
        },
        {
            "tema": "Por que esto pertenece a Fase 5",
            "explicacion_didactica": (
                "El pipeline profesional ubica limpieza, feature engineering, codificacion categorica y normalizacion en Data Preparation."
            ),
            "decision_fase5": "Fase 5 deja los datos listos; Fase 6 modela.",
            "donde_ver_detalle": "CONTEXTOS/pipeline_datascience.md, Paso 3",
        },
    ]
    return pd.DataFrame(rows)


def build_guide_sheet(
    feature_matrix: pd.DataFrame,
    coverage_report: pd.DataFrame,
    transform_params: pd.DataFrame | None,
) -> pd.DataFrame:
    countries = build_countries_sheet(feature_matrix)
    n_observed = len(get_mvp_variables())
    income_counts = countries["grupo_ingreso"].fillna("Sin clasificacion").value_counts().to_dict()
    income_text = "; ".join(f"{k}: {v}" for k, v in income_counts.items())
    rows = [
        ("1. Que es este Excel", "Este workbook corresponde a Fase 5 v2.1+ del proyecto Research_AI_law. La Fase 5 es una fase de preparacion auditable y contractual, no una fase de pre-modelado predictivo. La muestra primaria contiene 43 paises preregistrados y 46 variables observadas reales. Particiones predictivas y conjuntos artificiales han sido eliminados."),
        ("2. Que NO es", "No es todavia el resultado de los modelos ni una prueba causal. La inferencia ocurre en Fase 6."),
        ("3. Hipotesis", "La hipotesis y subpreguntas estan en 1_Hipotesis."),
        ("4. Hoja mas importante para lectura sustantiva", f"6_Matriz_40_Humana (nombre legacy): 43 paises x {n_observed} variables observadas. Esta es la matriz que debe revisar una autoridad."),
        ("5. Por que tambien existe una matriz tecnica", "11_Features_Fase6 y Muestra_Analitica_v2_1 son para el equipo tecnico: incluyen transformaciones, agregados y flags de muestra analitica; el uso de datos es observacional completo."),
        ("6. Donde estan exactamente las 46 variables", f"5_Variables_40 (legacy) lista V01-V{n_observed:02d}; 5b_Variables_46_Detalle destaca las 6 nuevas de v2.0."),
        ("7. Como auditar una celda", f"Use 2_Como_Auditar: elegir variable V01-V{n_observed:02d}, ir a 6_Matriz_40_Humana y luego filtrar 13_Trazabilidad."),
        ("8. Que significan los colores", "7_Leyenda_Colores explica cada color exacto. En 6_Matriz_40_Humana los colores son por rol analitico, no todos amarillos."),
        ("9. Que significa income_group", "Es grupo de ingreso heredado de Fase 3 / World Bank. La definicion y conteos estan en 4_Ingreso_Region."),
        ("10. Por que Taiwan tiene vacios", "Taiwan fue incluido por relevancia IA, pero Fase 3 no trae region/income_group comparable y varias fuentes WB/WIPO no aplican. Ver 8_Casos_Atencion."),
        ("11. Normalizacion", "9_Normalizacion explica por que se normaliza, que se hizo, que no corrige y como se audita."),
        ("12. Vacio no significa cero", "Fase 5 no imputa. Una celda vacia significa dato no observado/no comparable en fuente."),
        ("13. Cobertura", f"La cobertura minima de las {n_observed} variables es {coverage_report['pct_complete'].min():.2f}%, sobre el umbral de 30%."),
        ("14. Grupos de ingreso presentes", income_text),
        ("15. Transformaciones", f"Las transformaciones derivadas con status zero_mad_or_not_estimable o similares se excluyen de modelado primario. Hay {0 if transform_params is None else len(transform_params)} registros en 14_Transformaciones."),
    ]
    return pd.DataFrame(rows, columns=["pregunta_de_lectura", "respuesta_clara"])


def _style_workbook(path, feature_matrix: pd.DataFrame) -> None:
    workbook = None
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = True
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
            cell.border = border
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            header = _safe_text(column_cells[0].value)
            max_len = min(max(len(_safe_text(c.value)) for c in column_cells[:200]) + 2, 55)
            if header in {
                "respuesta_clara", "explicacion_didactica", "rol_en_el_proyecto",
                "motivo_inclusion_mvp", "lectura_humana", "known_limitations",
                "que_debe_quedar_claro", "nota_para_revision", "detalle",
                "que_significa_transform", "donde_auditar_en_excel",
                "decision_fase5", "donde_ver_detalle",
            }:
                max_len = 65
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len)

    catalog = _with_variable_ids(build_variable_catalog())
    role_by_var = dict(zip(catalog["variable_matriz"], catalog["rol_mvp"]))
    label_to_var = dict(zip(
        catalog["id_variable"] + " - " + catalog["nombre_corto_para_excel"],
        catalog["variable_matriz"],
    ))

    if "6_Matriz_40_Humana" in workbook.sheetnames:
        ws = workbook["6_Matriz_40_Humana"]
        for cell in ws[1]:
            header = _safe_text(cell.value)
            original_header = {
                "codigo_iso3": "iso3",
                "pais": "country_name_canonical",
                "region_geografica": "region",
                "grupo_ingreso": "income_group",
            }.get(header, label_to_var.get(header, header))
            if original_header in get_mvp_variables():
                color = ROLE_COLORS.get(role_by_var.get(original_header, ""), "FFF2CC")
            else:
                color = COLUMN_TYPE_COLORS["metadato"]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    if "11_Features_Fase6" in workbook.sheetnames:
        ws = workbook["11_Features_Fase6"]
        col_dict = build_column_dictionary(feature_matrix).set_index("columna")["tipo_columna"].to_dict()
        for cell in ws[1]:
            col_type = col_dict.get(_safe_text(cell.value), "auditoria")
            color = COLUMN_TYPE_COLORS.get(col_type, "EFEFEF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    if "12_Diccionario_Cols" in workbook.sheetnames:
        ws = workbook["12_Diccionario_Cols"]
        type_col_idx = None
        for cell in ws[1]:
            if cell.value == "tipo_columna":
                type_col_idx = cell.column
        if type_col_idx:
            for row in range(2, ws.max_row + 1):
                col_type = ws.cell(row=row, column=type_col_idx).value
                color = COLUMN_TYPE_COLORS.get(_safe_text(col_type), "EFEFEF")
                for cell in ws[row]:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    if "7_Leyenda_Colores" in workbook.sheetnames:
        ws = workbook["7_Leyenda_Colores"]
        color_col = None
        for cell in ws[1]:
            if cell.value == "color_hex":
                color_col = cell.column
        if color_col:
            for row in range(2, ws.max_row + 1):
                color = _safe_text(ws.cell(row=row, column=color_col).value)
                if color:
                    ws.cell(row=row, column=color_col).fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

    if "8_Casos_Atencion" in workbook.sheetnames:
        ws = workbook["8_Casos_Atencion"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "TWN":
                for cell in ws[row]:
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    tab_colors = {
        "0_Leer_Primero": "4472C4",
        "1_Hipotesis": "4472C4",
        "2_Como_Auditar": "70AD47",
        "3_Paises_43": "70AD47",
        "4_Ingreso_Region": "70AD47",
        "5_Variables_40": "FFC000",
        "5b_Variables_46_Detalle": "FFC000",
        "6_Matriz_40_Humana": "FFC000",
        "7_Leyenda_Colores": "A5A5A5",
        "8_Casos_Atencion": "C00000",
        "9_Normalizacion": "70AD47",
        "10_Cobertura": "5B9BD5",
        "11_Features_Fase6": "C55A11",
        "Muestra_Analitica_v2_1": "C55A11",
        "12_Diccionario_Cols": "A5A5A5",
        "13_Trazabilidad": "8064A2",
        "14_Transformaciones": "70AD47",
    }
    for name, color in tab_colors.items():
        if name in workbook.sheetnames:
            workbook[name].sheet_properties.tabColor = color
    workbook.save(path)


def write_audit_excel(
    wide_mvp: pd.DataFrame,
    feature_matrix: pd.DataFrame,
    coverage_report: pd.DataFrame,
    transform_params: pd.DataFrame | None = None,
    output_path=None,
    membership: pd.DataFrame | None = None,
) -> str:
    if output_path is None:
        output_path = FASE5_OUTPUTS / "MVP_AUDITABLE.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if transform_params is None:
        transform_params = pd.DataFrame()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        build_guide_sheet(feature_matrix, coverage_report, transform_params).to_excel(
            writer, sheet_name="0_Leer_Primero", index=False
        )
        build_hypothesis_sheet().to_excel(writer, sheet_name="1_Hipotesis", index=False)
        build_audit_protocol_sheet().to_excel(writer, sheet_name="2_Como_Auditar", index=False)
        build_countries_sheet(wide_mvp).to_excel(writer, sheet_name="3_Paises_43", index=False)
        build_income_region_sheet(feature_matrix).to_excel(writer, sheet_name="4_Ingreso_Region", index=False)
        build_variables_sheet().to_excel(writer, sheet_name="5_Variables_40", index=False)
        build_v2_variables_detail_sheet().to_excel(writer, sheet_name="5b_Variables_46_Detalle", index=False)
        build_observed_matrix_sheet(feature_matrix).to_excel(writer, sheet_name="6_Matriz_40_Humana", index=False)
        build_color_legend_sheet().to_excel(writer, sheet_name="7_Leyenda_Colores", index=False)
        build_data_attention_sheet(feature_matrix).to_excel(writer, sheet_name="8_Casos_Atencion", index=False)
        build_normalization_sheet(transform_params).to_excel(writer, sheet_name="9_Normalizacion", index=False)
        coverage_report.to_excel(writer, sheet_name="10_Cobertura", index=False)
        feature_matrix.to_excel(writer, sheet_name="11_Features_Fase6", index=False)
        if membership is not None:
            membership.to_excel(writer, sheet_name="Muestra_Analitica_v2_1", index=False)
        build_column_dictionary(feature_matrix).to_excel(writer, sheet_name="12_Diccionario_Cols", index=False)
        build_traceability_sheet().to_excel(writer, sheet_name="13_Trazabilidad", index=False)
        transform_params.to_excel(writer, sheet_name="14_Transformaciones", index=False)

    _style_workbook(output_path, feature_matrix)
    return str(output_path)
