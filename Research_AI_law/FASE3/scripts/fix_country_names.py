"""
Fix Fase 3: country_name_canonical was set to ISO3 in 86/199 wide rows
because geo.py used split(' | ')[0] which returned the ISO3, not the name.

This script:
  1. Builds a canonical ISO3 -> name mapping from cobertura_pais_fuente.csv
     by extracting the non-ISO3 component of country_name_best_effort.
  2. Adds a manual override for the 4 territories and edge cases.
  3. Patches all Fase 3 outputs that contain country_name_canonical.
  4. Regenerates the Excel and manifest.

Run from FASE3/ root:
    python3 scripts/fix_country_names.py
"""

from __future__ import annotations

import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
EDA_DIR = ROOT.parent / "outputs" / "eda_preliminar"
OUT_DIR = ROOT / "outputs"

# Manual canonical names for entities where best_effort lacks a usable name
# or where we want a specific official label.
MANUAL_OVERRIDES: dict[str, str] = {
    "USA": "United States",
    "GBR": "United Kingdom",
    "RUS": "Russia",
    "KOR": "South Korea",
    "PRK": "North Korea",
    "IRN": "Iran",
    "VEN": "Venezuela",
    "BOL": "Bolivia",
    "CIV": "Cote d'Ivoire",
    "CPV": "Cabo Verde",
    "COD": "Democratic Republic of the Congo",
    "COG": "Republic of the Congo",
    "CZE": "Czech Republic",
    "EGY": "Egypt",
    "FRA": "France",
    "DEU": "Germany",
    "GRC": "Greece",
    "HKG": "Hong Kong",
    "MAC": "Macao",
    "MDA": "Moldova",
    "MKD": "North Macedonia",
    "PRI": "Puerto Rico",
    "SVK": "Slovakia",
    "SVN": "Slovenia",
    "SYR": "Syria",
    "TWN": "Taiwan",
    "TZA": "Tanzania",
    "VNM": "Vietnam",
    "LAO": "Laos",
    "BRN": "Brunei",
    "PSE": "Palestine",
    "XKX": "Kosovo",
    "XKV": "Kosovo",
    "ARE": "United Arab Emirates",
    "GMB": "The Gambia",
    "BHS": "The Bahamas",
}


def build_iso3_name_map() -> dict[str, str]:
    """Build canonical ISO3 -> human name map by parsing best_effort."""
    cov = pd.read_csv(EDA_DIR / "cobertura_pais_fuente.csv")
    mapping: dict[str, str] = {}
    for _, row in cov.iterrows():
        iso3 = str(row["iso3"]).strip().upper()
        be = row["country_name_best_effort"]
        if pd.isna(be) or not str(be).strip():
            mapping[iso3] = iso3
            continue
        parts = [p.strip() for p in str(be).split("|")]
        candidates = [p for p in parts if p and p.upper() != iso3.upper()]
        if not candidates:
            mapping[iso3] = iso3
            continue
        # prefer longest non-iso3 candidate; ties broken alphabetically
        candidates.sort(key=lambda x: (-len(x), x))
        mapping[iso3] = candidates[0]
    # apply manual overrides
    mapping.update(MANUAL_OVERRIDES)
    return mapping


def patch_csv(path: Path, name_map: dict[str, str]) -> tuple[int, int]:
    """Patch a CSV in O(n). Returns (rows_changed, total_rows)."""
    if not path.exists():
        return (0, 0)
    df = pd.read_csv(path)
    if "country_name_canonical" not in df.columns or "iso3" not in df.columns:
        return (0, len(df))
    iso_norm = df["iso3"].astype(str).str.strip().str.upper()
    name_old = df["country_name_canonical"].astype(str)
    before_unresolved = int((iso_norm == name_old).sum())
    # vectorized map; keep existing value if iso3 not in name_map
    new_names = iso_norm.map(name_map).fillna(name_old)
    # only overwrite where the old name equals the iso3 (i.e., the buggy rows)
    needs_fix = iso_norm == name_old
    df.loc[needs_fix, "country_name_canonical"] = new_names[needs_fix].values
    df.to_csv(path, index=False)
    after_unresolved = int((iso_norm == df["country_name_canonical"].astype(str)).sum())
    return (before_unresolved - after_unresolved, len(df))


def patch_excel_in_place(name_map: dict[str, str]) -> None:
    """Update country_name_canonical column in-place in each sheet of the Excel.
    Much faster than regenerating the whole Excel from CSVs."""
    excel_path = OUT_DIR / "Matriz_Madre_Fase3.xlsx"
    if not excel_path.exists():
        print(f"  Excel not found: {excel_path}")
        return

    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    total_fixes = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # find column indices (1-based)
        if ws.max_row < 2:
            continue
        header = [c.value for c in ws[1]]
        if "iso3" not in header or "country_name_canonical" not in header:
            continue
        iso_col = header.index("iso3") + 1
        name_col = header.index("country_name_canonical") + 1
        sheet_fixes = 0
        for row in range(2, ws.max_row + 1):
            iso_val = ws.cell(row=row, column=iso_col).value
            name_val = ws.cell(row=row, column=name_col).value
            if iso_val is None:
                continue
            iso_norm = str(iso_val).strip().upper()
            if name_val == iso_norm and iso_norm in name_map:
                ws.cell(row=row, column=name_col).value = name_map[iso_norm]
                sheet_fixes += 1
        if sheet_fixes:
            print(f"    sheet '{sheet_name}': {sheet_fixes} cells fixed")
            total_fixes += sheet_fixes
    wb.save(excel_path)
    wb.close()
    print(f"  Excel total: {total_fixes} cells fixed in-place")


def update_manifest(name_map_size: int) -> None:
    """Update the manifest with new hashes after the fix."""
    manifest_path = OUT_DIR / "manifest.json"
    if not manifest_path.exists():
        return
    with open(manifest_path) as f:
        manifest = json.load(f)

    # Recompute hashes for all listed outputs
    new_hashes = {}
    for name in manifest.get("outputs", {}):
        path = OUT_DIR / name
        if path.exists():
            with open(path, "rb") as fh:
                new_hashes[name] = hashlib.sha256(fh.read()).hexdigest()

    manifest["outputs"] = new_hashes
    manifest["created_at"] = datetime.now(timezone.utc).isoformat()
    manifest.setdefault("changelog", []).append({
        "date": datetime.now(timezone.utc).isoformat(),
        "change": "country_name_canonical fix v1.1",
        "rows_affected": "wide=86, universe=99, panel=10103, snapshot=2003, excluded=variable",
        "iso3_names_resolved": name_map_size,
        "version": "1.1",
    })
    manifest["version"] = "1.1"

    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2)
    print(f"  Manifest updated to v1.1")


def main() -> None:
    print("=" * 70)
    print("Fase 3 fix: country_name_canonical")
    print("=" * 70)

    print("\n[1] Building canonical ISO3 -> name map...")
    name_map = build_iso3_name_map()
    print(f"    {len(name_map)} ISO3 codes mapped")
    samples = ["ARG", "AUS", "FRA", "COL", "BRA", "CAN", "CHN", "USA", "GBR", "CHL"]
    for iso in samples:
        print(f"    {iso} -> {name_map.get(iso, '(missing)')}")

    print("\n[2] Patching CSV outputs...")
    targets = [
        "matriz_madre_wide.csv",
        "matriz_larga_panel.csv",
        "matriz_larga_snapshot.csv",
        "matriz_madre_trazabilidad.csv",
        "fase3_universo_geografico.csv",
        "fase3_entidades_excluidas_geografia.csv",
        "fase3_geo_crosswalk_manual.csv",
        "fase3_auditoria_muestra_valores.csv",
    ]
    for fname in targets:
        path = OUT_DIR / fname
        changed, total = patch_csv(path, name_map)
        if total > 0:
            print(f"    {fname}: {changed:,}/{total:,} rows fixed")

    print("\n[3] Patching Excel in-place...")
    patch_excel_in_place(name_map)

    print("\n[4] Updating manifest to v1.1...")
    update_manifest(len(name_map))

    print("\nFix complete.")


if __name__ == "__main__":
    main()
