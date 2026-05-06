from openpyxl import load_workbook

from fase3_pipeline.config import EXCEL_SHEETS, OUTPUT_FILES


def test_all_required_outputs_exist(output_dir):
    missing = [name for name in OUTPUT_FILES if not (output_dir / name).exists()]
    assert not missing


def test_excel_has_13_complete_sheets(output_dir, panel, snapshot, wide, dictionary, traceability):
    wb = load_workbook(output_dir / "Matriz_Madre_Fase3.xlsx", read_only=True)
    assert wb.sheetnames == EXCEL_SHEETS
    assert wb["Matriz Madre"].max_row == len(wide) + 1
    assert wb["Matriz Madre"].max_column == len(wide.columns)
    assert wb["Matriz Larga Panel"].max_row == len(panel) + 1
    assert wb["Matriz Larga Snapshot"].max_row == len(snapshot) + 1
    assert wb["Diccionario Variables"].max_row == len(dictionary) + 1
    assert wb["Trazabilidad"].max_row == len(traceability) + 1
