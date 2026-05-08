"""Tarea 8: Extraer Bloque Gobernanza (WGI) desde Excel.
El WGI NO esta disponible via API v2. Se descarga el Excel oficial.
Fuente: https://www.worldbank.org/content/dam/sites/govindicators/doc/wgidataset_with_sourcedata-2025.xlsx
"""
import openpyxl
import csv
import os
from datetime import datetime

EXCEL_PATH = "/Users/francoia/Documents/Research_AI_law/World Bank WDI/data_v2/raw/wgidataset_with_sourcedata-2025.xlsx"
OUTPUT_DIR = "/Users/francoia/Documents/Research_AI_law/World Bank WDI/data_v2/raw"
AUDIT_FILE = os.path.join(OUTPUT_DIR, "tarea8_audit_gobernanza.csv")
COUNTRIES_FILE = "/Users/francoia/Documents/Research_AI_law/World Bank WDI/data_v2/raw/tarea2_paises_metadata.csv"

# Sheet code -> canonical name mapping
SHEET_MAP = {
    "va": "voice_accountability",
    "pv": "political_stability",
    "ge": "government_effectiveness",
    "rq": "regulatory_quality",
    "rl": "rule_of_law",
    "cc": "control_of_corruption",
}

# WGI indicator codes for canonical names
WGI_CODES = {
    "voice_accountability":       "VA.EST",
    "political_stability":        "PV.EST",
    "government_effectiveness":   "GE.EST",
    "regulatory_quality":         "RQ.EST",
    "rule_of_law":                "RL.EST",
    "control_of_corruption":      "CC.EST",
}

def load_valid_countries():
    valid = set()
    with open(COUNTRIES_FILE, "r") as f:
        for row in csv.DictReader(f):
            valid.add(row["iso3"])
    return valid

def main():
    print("=" * 80)
    print("TAREA 8: Extraccion Bloque Gobernanza (WGI via Excel)")
    print(f"Hora: {datetime.now().isoformat()}")
    print(f"Fuente: govindicators/doc/wgidataset_with_sourcedata-2025.xlsx")
    print("=" * 80)
    
    valid_iso3 = load_valid_countries()
    print(f"Valid countries: {len(valid_iso3)}")
    
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    print(f"Excel sheets: {wb.sheetnames}")
    
    all_audit = []
    total_rows_est = 0
    total_rows_se = 0
    year_range = range(2018, 2025)  # 2018-2024
    
    for sheet_code, canonical in SHEET_MAP.items():
        if sheet_code not in wb.sheetnames:
            print(f"  SKIP: {sheet_code} not found in Excel")
            continue
        
        ws = wb[sheet_code]
        print(f"\n[Sheet: {sheet_code} -> {canonical}]")
        
        rows_est = []   # Estimate rows
        rows_se = []    # Standard Error rows
        n_skipped = {"not_country": 0, "bad_year": 0, "null_value": 0}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            iso3 = row[2]  # col C: Economy code
            year = row[5]  # col F: Year
            est_val = row[8]  # col I: Governance estimate
            se_val = row[9]   # col J: Standard error
            country_name = row[1]  # col B: Economy name
            
            if not iso3 or len(str(iso3)) != 3:
                n_skipped["not_country"] += 1
                continue
            
            if year not in year_range:
                n_skipped["bad_year"] += 1
                continue
            
            # Estimate
            if est_val is not None:
                rows_est.append({
                    "iso3": iso3,
                    "year": year,
                    "country_name": country_name,
                    "value": est_val,
                    "wb_indicator": WGI_CODES[canonical],
                    "canonical_name": canonical,
                })
            else:
                n_skipped["null_value"] += 1
            
            # Standard Error
            if se_val is not None:
                rows_se.append({
                    "iso3": iso3,
                    "year": year,
                    "country_name": country_name,
                    "value": se_val,
                    "wb_indicator": WGI_CODES[canonical].replace(".EST", ".STD.ERR"),
                    "canonical_name": f"{canonical}_se",
                })
        
        # Save Estimate CSV
        est_path = os.path.join(OUTPUT_DIR, f"tarea8_{canonical}.csv")
        with open(est_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["iso3","year","country_name","value","wb_indicator","canonical_name"])
            w.writeheader()
            w.writerows(rows_est)
        
        # Save SE CSV
        se_path = os.path.join(OUTPUT_DIR, f"tarea8_{canonical}_se.csv")
        with open(se_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["iso3","year","country_name","value","wb_indicator","canonical_name"])
            w.writeheader()
            w.writerows(rows_se)
        
        years_est = sorted(set(r["year"] for r in rows_est))
        countries_est = sorted(set(r["iso3"] for r in rows_est))
        
        print(f"  Estimate:  {len(rows_est)} rows | {len(countries_est)} countries | years: {years_est}")
        print(f"  Std.Error: {len(rows_se)} rows")
        print(f"  Skipped:   {n_skipped}")
        
        # Audit entry for estimate
        all_audit.append({
            "bloque": "Gobernanza",
            "canonical_name": canonical,
            "wb_code": WGI_CODES[canonical],
            "wb_name": f"WGI {canonical} (Estimate)",
            "last_updated_wb": "2025 WGI update",
            "n_rows_final": len(rows_est),
            "n_countries": len(countries_est),
            "years_data": ", ".join(str(y) for y in years_est),
            "output_file": est_path,
            "extraction_timestamp": datetime.now().isoformat(),
            "source": "WGI Excel: wgidataset_with_sourcedata-2025.xlsx",
        })
        
        # Audit entry for SE
        years_se = sorted(set(r["year"] for r in rows_se))
        countries_se = sorted(set(r["iso3"] for r in rows_se))
        all_audit.append({
            "bloque": "Gobernanza",
            "canonical_name": f"{canonical}_se",
            "wb_code": WGI_CODES[canonical].replace(".EST", ".STD.ERR"),
            "wb_name": f"WGI {canonical} (Standard Error)",
            "last_updated_wb": "2025 WGI update",
            "n_rows_final": len(rows_se),
            "n_countries": len(countries_se),
            "years_data": ", ".join(str(y) for y in years_se),
            "output_file": se_path,
            "extraction_timestamp": datetime.now().isoformat(),
            "source": "WGI Excel: wgidataset_with_sourcedata-2025.xlsx",
        })
        
        total_rows_est += len(rows_est)
        total_rows_se += len(rows_se)
    
    wb.close()
    
    # Save audit
    fieldnames = [
        "bloque","canonical_name","wb_code","wb_name","last_updated_wb",
        "n_rows_final","n_countries","years_data","output_file",
        "extraction_timestamp","source"
    ]
    with open(AUDIT_FILE, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(all_audit)
    
    print(f"\n{'='*80}")
    print(f"TAREA 8 COMPLETADA")
    print(f"  Indicators: 6 WGI + 6 SE = 12 extraidos")
    print(f"  Estimate rows: {total_rows_est}")
    print(f"  SE rows:       {total_rows_se}")
    print(f"  Audit log:     {AUDIT_FILE}")

if __name__ == "__main__":
    main()
