"""Tareas 11-13: Generar Excel final WB_WDI_WGI_unificado.xlsx con 10 pestañas.

Estructura:
1. MATRIZ_COMPLETA     - Matriz ancha unificada
2. BLOQUE_DESARROLLO   - 7 variables desarrollo economico
3. BLOQUE_APERTURA     - 5 variables apertura/inversion
4. BLOQUE_DIGITAL      - 5 variables infraestructura digital
5. BLOQUE_CAPITAL_HUMANO - 4 variables capital humano
6. BLOQUE_INNOVACION   - 4 variables innovacion
7. BLOQUE_GOBERNANZA   - 12 variables gobernanza WGI
8. DICCIONARIO         - Metadata completa de cada variable
9. AUDITORIA           - Log de extraccion con trazabilidad
10. FUENTES             - Referencias oficiales
"""
import pandas as pd
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = "/Users/francoia/Documents/Research_AI_law/World Bank WDI"
DATA_DIR = os.path.join(BASE_DIR, "data_v2")
OUTPUT = os.path.join(BASE_DIR, "WB_WDI_WGI_unificado.xlsx")

MATRIZ_FILE = os.path.join(DATA_DIR, "MATRIZ_COMPLETA.csv")
DICCIONARIO_FILE = os.path.join(DATA_DIR, "DICCIONARIO.csv")
AUDIT_FILES = os.path.join(DATA_DIR, "raw")

# Which variable columns belong to each block
BLOCK_VARS = {
    "Desarrollo": [
        "gdp_per_capita_ppp", "gdp_current_usd", "gdp_growth_annual_pct",
        "inflation_consumer_prices", "unemployment_rate", "population", "labor_force"
    ],
    "Apertura": [
        "exports_pct_gdp", "trade_pct_gdp", "fdi_net_inflows",
        "fdi_pct_gdp", "gross_capital_formation_pct_gdp"
    ],
    "Digital": [
        "internet_penetration", "mobile_subscriptions_per100", "fixed_broadband_per100",
        "secure_servers_per_1m", "electric_consumption_kwh_pc"
    ],
    "Capital_Humano": [
        "tertiary_education_enrollment", "education_expenditure_pct_gdp",
        "rd_expenditure_pct_gdp", "researchers_rd_per_million"
    ],
    "Innovacion": [
        "patent_applications_residents", "high_tech_exports_pct_manufactured",
        "ict_goods_exports_pct", "ict_service_exports_pct"
    ],
    "Gobernanza": [
        "control_of_corruption", "government_effectiveness", "political_stability",
        "regulatory_quality", "rule_of_law", "voice_accountability",
        "control_of_corruption_se", "government_effectiveness_se", "political_stability_se",
        "regulatory_quality_se", "rule_of_law_se", "voice_accountability_se"
    ],
}

# Styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def freeze_first_row(ws):
    ws.freeze_panes = "A2"

def auto_width(ws, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:100]:  # sample first 100 rows
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max_len + 3, max_width)
        ws.column_dimensions[col_letter].width = adjusted

def get_audit_data():
    """Compile audit records from raw audit CSVs."""
    audit_files = [
        os.path.join(AUDIT_FILES, "tarea3_audit_desarrollo.csv"),
        os.path.join(AUDIT_FILES, "tarea47_audit_Apertura.csv"),
        os.path.join(AUDIT_FILES, "tarea47_audit_Digital.csv"),
        os.path.join(AUDIT_FILES, "tarea47_audit_Capital_Humano.csv"),
        os.path.join(AUDIT_FILES, "tarea47_audit_Innovacion.csv"),
        os.path.join(AUDIT_FILES, "tarea8_audit_gobernanza.csv"),
    ]
    
    rows = []
    for f in audit_files:
        if os.path.exists(f):
            with open(f, "r") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    rows.append(row)
    return rows

def main():
    print("=" * 80)
    print("TAREAS 11-13: Generar Excel WB_WDI_WGI_unificado.xlsx")
    print(f"Hora: {datetime.now().isoformat()}")
    print("=" * 80)
    
    # Load data
    print("\n1. Cargando MATRIZ_COMPLETA...")
    df = pd.read_csv(MATRIZ_FILE)
    meta_cols = ["iso3", "country_name", "wb_region", "wb_income_group"]
    print(f"   {df.shape[0]} filas, {df.shape[1]} columnas")
    
    print("2. Cargando DICCIONARIO...")
    df_dict = pd.read_csv(DICCIONARIO_FILE)
    print(f"   {df_dict.shape[0]} variables")
    
    print("3. Compilando AUDITORIA...")
    audit_data = get_audit_data()
    print(f"   {len(audit_data)} registros de auditoria")
    
    # Create workbook
    print("\n4. Creando Excel con 10 pestañas...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # ---- SHEET 1: MATRIZ_COMPLETA ----
    print("   Sheet 1/10: MATRIZ_COMPLETA")
    ws1 = wb.create_sheet("MATRIZ_COMPLETA")
    for col_idx, col_name in enumerate(df.columns, 1):
        ws1.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws1.cell(row=row_idx, column=col_idx, value=val)
    style_header(ws1, len(df.columns))
    freeze_first_row(ws1)
    
    # ---- SHEETS 2-7: Block sheets ----
    for block_name, var_list in BLOCK_VARS.items():
        sheet_name = f"BLOQUE_{block_name.upper()}"
        print(f"   Sheet: {sheet_name}")
        ws = wb.create_sheet(sheet_name)
        
        # Get columns for this block
        block_cols = meta_cols.copy()
        for var in var_list:
            matching = [c for c in df.columns if c.startswith(var + "_")]
            block_cols.extend(sorted(matching))
        # Only include columns that exist
        block_cols = [c for c in block_cols if c in df.columns]
        
        # Write data
        for col_idx, col_name in enumerate(block_cols, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, col_name in enumerate(block_cols, 1):
                ws.cell(row=row_idx, column=col_idx, value=getattr(row, col_name))
        
        style_header(ws, len(block_cols))
        freeze_first_row(ws)
    
    # ---- SHEET 8: DICCIONARIO ----
    print("   Sheet 8/10: DICCIONARIO")
    ws8 = wb.create_sheet("DICCIONARIO")
    for col_idx, col_name in enumerate(df_dict.columns, 1):
        ws8.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df_dict.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws8.cell(row=row_idx, column=col_idx, value=val)
    style_header(ws8, len(df_dict.columns))
    freeze_first_row(ws8)
    auto_width(ws8, max_width=80)
    
    # ---- SHEET 9: AUDITORIA ----
    print("   Sheet 9/10: AUDITORIA")
    ws9 = wb.create_sheet("AUDITORIA")
    # Compile unique fieldnames from audit data
    audit_fieldnames = [
        "bloque", "canonical_name", "wb_code", "wb_name", "last_updated_wb",
        "n_rows_final", "n_countries", "years_data", "output_file",
        "extraction_timestamp", "api_url", "source", "n_aggregates_filtered",
        "n_null_filtered", "total_records_api"
    ]
    # Filter to fields that exist in audit data
    if audit_data:
        available_fields = [f for f in audit_fieldnames if f in audit_data[0] or f == "source"]
        available_fields = [f for f in audit_fieldnames if any(f in d for d in audit_data)]
    else:
        available_fields = audit_fieldnames
    
    # Re-read actual fields from first record
    actual_fields = []
    if audit_data:
        actual_fields = list(audit_data[0].keys())
    
    for col_idx, col_name in enumerate(actual_fields, 1):
        ws9.cell(row=1, column=col_idx, value=col_name)
    for row_idx, record in enumerate(audit_data, 2):
        for col_idx, col_name in enumerate(actual_fields, 1):
            ws9.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
    style_header(ws9, len(actual_fields))
    freeze_first_row(ws9)
    auto_width(ws9, max_width=80)
    
    # ---- SHEET 10: FUENTES ----
    print("   Sheet 10/10: FUENTES")
    ws10 = wb.create_sheet("FUENTES")
    fuentes = [
        ["Fuente", "Tipo", "URL", "Descripcion", "Fecha Acceso"],
        ["World Bank WDI API v2", "API REST", "https://api.worldbank.org/v2/country/all/indicator/{CODE}?date=2018:2025&format=json&source=2", "World Development Indicators. 25 indicadores extraidos via API programatica. source_id=2. Max per_page=20000. Filtrado: solo paises reales (excluye agregados regionales).", datetime.now().strftime("%Y-%m-%d")],
        ["World Bank DataBank (WDI)", "Web", "https://databank.worldbank.org/source/world-development-indicators", "Interfaz web del Banco Mundial para consultar WDI. Permite verificacion manual de cualquier indicador por pais y año.", datetime.now().strftime("%Y-%m-%d")],
        ["World Bank WGI (Excel)", "Excel", "https://www.worldbank.org/content/dam/sites/govindicators/doc/wgidataset_with_sourcedata-2025.xlsx", "Worldwide Governance Indicators, 2025 update. Datos 1996-2024. 6 dimensiones de gobernanza + errores estandar. NO disponible via API v2; descarga directa de Excel.", datetime.now().strftime("%Y-%m-%d")],
        ["World Bank WGI (Web)", "Web", "https://www.worldbank.org/en/publication/worldwide-governance-indicators", "Pagina oficial WGI. Documentacion metodologica, FAQs, y acceso interactive a datos.", datetime.now().strftime("%Y-%m-%d")],
        ["World Bank Country API", "API REST", "https://api.worldbank.org/v2/country?format=json&per_page=500", "Metadata de paises: iso3, nombre, region WB (7 categorias), nivel de ingreso WB (4 categorias), capital, coordenadas.", datetime.now().strftime("%Y-%m-%d")],
        ["World Bank Indicator Metadata", "API REST", "https://api.worldbank.org/v2/indicator/{CODE}?format=json", "Metadata de cada indicador: nombre, fuente, organizacion, ultima actualizacion.", datetime.now().strftime("%Y-%m-%d")],
        ["World Bank Data Help Desk", "Web", "https://datahelpdesk.worldbank.org/knowledgebase/articles/898599-indicator-api-queries", "Documentacion oficial de la API del Banco Mundial: parametros, limites, buenas practicas.", datetime.now().strftime("%Y-%m-%d")],
        ["", "", "", "", ""],
        ["NOTA METODOLOGICA", "", "", "Todos los datos en este Excel provienen EXCLUSIVAMENTE de las fuentes arriba indicadas. No se realizaron imputaciones, interpolaciones ni estimaciones propias. Los valores son exactamente los devueltos por las fuentes del Banco Mundial. Celdas vacias = dato no disponible en la fuente.", ""],
        ["", "", "", "", ""],
        ["FORMATO DE VERIFICACION", "", "", "Para auditar un valor: (1) Identificar el codigo WB en la pestaña DICCIONARIO, (2) Usar URL web con el codigo de pais ISO3, (3) Comparar el valor en pantalla con el valor en este Excel.", ""],
        ["Ejemplo", "", "https://data.worldbank.org/indicator/NY.GDP.PCAP.PP.CD?locations=CL", "Verificar PIB per capita PPP de Chile (CHL).", ""],
    ]
    for row_idx, row_data in enumerate(fuentes, 1):
        for col_idx, val in enumerate(row_data, 1):
            ws10.cell(row=row_idx, column=col_idx, value=val)
    # Style header
    for col in range(1, 5 + 1):
        cell = ws10.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    freeze_first_row(ws10)
    auto_width(ws10, max_width=100)
    
    # Save
    print(f"\n5. Guardando: {OUTPUT}")
    wb.save(OUTPUT)
    
    # Summary
    import os as _os
    size_mb = _os.path.getsize(OUTPUT) / (1024 * 1024)
    print(f"\n{'='*80}")
    print(f"EXCEL GENERADO EXITOSAMENTE")
    print(f"  Archivo: {OUTPUT}")
    print(f"  Tamaño:  {size_mb:.1f} MB")
    print(f"  Pestañas: 10")
    print(f"  Filas (MATRIZ): {df.shape[0]} paises")
    print(f"  Columnas (MATRIZ): {df.shape[1]}")
    print(f"  Variables unicas: 37")
    print(f"  Fuentes: 100% World Bank (WDI API + WGI Excel)")
    print(f"  Datos imputados: 0 (solo datos reales)")
    print(f"  Fecha extraccion: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

if __name__ == "__main__":
    main()
