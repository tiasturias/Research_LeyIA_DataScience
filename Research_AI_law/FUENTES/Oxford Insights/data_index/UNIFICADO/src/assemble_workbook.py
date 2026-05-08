"""
assemble_workbook.py — Paso 10.0
Ensambla el workbook final Oxford_Insights_Unificado.xlsx con 9 hojas.
Output: UNIFICADO/Oxford_Insights_Unificado.xlsx
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import *
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define sheets: (sheet_name, source_csv, column mapping)
SHEETS = [
    ("Consolidado", "consolidado_final.csv", None),
    ("Detalle_2019", "detalle_2019.csv", None),
    ("Detalle_2020_2024", "detalle_2020_2024.csv", None),
    ("Detalle_2025", "detalle_2025.csv", None),
    ("Indicadores_2023", "indicadores_2023.csv", None),
    ("Rankings_Regionales_2019", "rankings_regionales_2019.csv", None),
    ("Diccionario_Variables", "diccionario_variables.csv", None),
    ("Fuentes_Directas", "fuentes_directas.csv", None),
    ("ISO3_Mapping", "iso3_mapping_sheet.csv", None),
]

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="top", wrap_text=False)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

for sheet_name, csv_name, _ in SHEETS:
    print(f"Escribiendo hoja: {sheet_name}...")
    df = pd.read_csv(output_path(csv_name), low_memory=False)
    
    ws = wb.create_sheet(title=sheet_name)
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows()):
        for col_idx, header in enumerate(headers):
            val = row[header]
            if pd.isna(val):
                cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=None)
            else:
                cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=val)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Auto-width (capped)
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        # Sample first 100 rows for max content width
        sample = df[header].astype(str).str.len().max() if len(df) > 0 else 0
        width = min(max(max_len, sample if pd.notna(sample) else 0) + 2, 50)
        ws.column_dimensions[col_letter].width = max(width, 10)
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"
    
    print(f"  {len(df)} filas x {len(headers)} columnas")

# Save
out_path = os.path.join(os.path.dirname(OUTPUT_DIR), "Oxford_Insights_Unificado.xlsx")
wb.save(out_path)
print(f"\nWorkbook guardado: {out_path}")
print("Hojas:")
for ws in wb.worksheets:
    print(f"  - {ws.title}: {ws.max_row} filas x {ws.max_column} columnas")
