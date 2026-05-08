"""
build_fuentes_v2.py — Fuentes_Directas con PRECISION ABSOLUTA
Cada variable mapeada a: archivo exacto, hoja exacta, columna exacta (letra+indice),
header exacto del Excel ORIGINAL, filas de datos, y notas de parsing.

Solo actualiza las hojas de documentacion (Fuentes_Directas, Diccionario_Variables),
NO toca las hojas de datos (Consolidado, Detalle_*, etc.).
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import *
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================================================================
# FUENTES_2019_DATA
# ================================================================
FUENTES_2019_DATA = [
    ("privacy_laws", "B", 1, "Privacy laws (yes = 1, no = 0)", "7-200",
     "Binario 0/1. Fila 1 (indice 1) = header. Col 13 (N) es gap."),
    ("ai_strategy", "C", 2, "AI strategy = 2 (forthcoming = 1, none = 0)\n", "7-200",
     "Ordinal 0/1/2. Header incluye salto de linea. Valor crudo."),
    ("data_availability", "D", 3, "Data availability", "7-200",
     "OKFN Open Data Index. Escala original 1-100."),
    ("govt_procurement_advanced_tech", "E", 4, "Gov't procurement of advanced technology products", "7-200",
     "Subindicator WEF Networked Readiness Index 2016."),
    ("data_capability_govt", "F", 5, "Data capability (in govt)", "7-200",
     "UN e-government index 2018. Escala 0-1."),
    ("technology_skills", "G", 6, "Technology skills", "7-200",
     "Subindicator WEF Global Competitiveness Report 2018."),
    ("ai_startups", "H", 7, "AI startups", "7-200",
     "Conteo Crunchbase. USA=5053, mayoria 0."),
    ("log_ai_startups", "I", 8, "Log of AI startups", "7-200",
     "Logaritmo de AI startups."),
    ("innovation_capability", "J", 9, "(Private sector) innovation capability", "7-200",
     "Subindicator WEF Global Competitiveness Report 2018."),
    ("digital_public_services", "K", 10, "Digital public services", "7-200",
     "UN online service index. Escala 0-1."),
    ("govt_effectiveness", "L", 11, "Effectiveness of government", "7-200",
     "World Bank 2017 Government Effectiveness."),
    ("ict_govt_vision", "M", 12, "Importance of ICTs to government vision of the future", "7-200",
     "Subindicator WEF Networked Readiness Index 2016."),
    ("norm_privacy_laws", "O", 14, "Privacy laws (yes = 1, no = 0)", "7-200",
     "NORMALISED. Version 0-1. Gap en col N(13)."),
    ("norm_ai_strategy", "P", 15, "AI strategy = 2 (forthcoming = 1, none = 0)\n", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_data_availability", "Q", 16, "Data availability", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_govt_procurement", "R", 17, "Gov't procurement of advanced technology products", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_data_capability", "S", 18, "Data capability (in govt)", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_technology_skills", "T", 19, "Technology skills", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_log_ai_startups", "U", 20, "Log of AI startups", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_innovation_capability", "V", 21, "(Private sector) innovation capability", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_digital_public_services", "W", 22, "Digital public services", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_govt_effectiveness", "X", 23, "Effectiveness of government", "7-200",
     "NORMALISED. Version 0-1."),
    ("norm_ict_govt_vision", "Y", 24, "Importance of ICTs to government vision of the future", "7-200",
     "NORMALISED. Version 0-1. Gap en col Z(25)."),
    ("avg_governance", "AD", 29, "Average: Governance", "7-200",
     "Promedio simple de Privacy laws + AI strategy normalizados."),
    ("avg_infrastructure_data", "AE", 30, "Average: Infrastructure and data", "7-200",
     "Promedio: Data availability + Gov't procurement + Data capability normalizados."),
    ("avg_skills_education", "AF", 31, "Average: Skills and education", "7-200",
     "Promedio: Technology skills + Log AI startups + Innovation capability normalizados."),
    ("avg_govt_public_services", "AG", 32, "Average: Government and public services", "7-200",
     "Promedio: Digital public services + Govt effectiveness + ICT govt vision normalizados."),
]

# ================================================================
# FUENTES_2019_RANKINGS
# ================================================================
REGIONS_2019 = {
    "global": (0, "A", "B", "C"),
    "asia_pacific": (3, "E", "F", "G"),
    "africa": (6, "I", "J", "K"),
    "latam": (9, "M", "N", "O"),
    "north_america": (12, "Q", "R", "S"),
    "eastern_europe": (15, "U", "V", "W"),
    "australasia": (18, "Y", "Z", "AA"),
    "western_europe": (21, "AC", "AD", "AE"),
}

FUENTES_2019_RANKINGS = []
for region, (cs, rl, cl, sl) in REGIONS_2019.items():
    rh = "Rank (of 194)" if region == "global" else "Rank (of 194 globally; regionally)"
    FUENTES_2019_RANKINGS.append(
        (f"rank_{region}", rl, cs, rh, "2-195",
         f"Region: {region}. Bloques de 3 cols (Rank, Country, Score) con gaps."))
    FUENTES_2019_RANKINGS.append(
        (f"score_{region}", sl, cs+2, "Score", "2-195",
         f"Region: {region}. Escala 0-10. Mismo score que INDEX SCORE."))
FUENTES_2019_RANKINGS.append(
    ("region_pertenece", "derivado", -1, "(derivado) Region de pertenencia", "2-195",
     "No es columna explicita. Determinada por la region donde aparece el pais."))

# ================================================================
# FUENTES_2020_2024
# ================================================================
YEAR_SHEET_MAP = {
    2020: ("2020-Government-AI-Readiness-Index-public-dataset.xlsx", "Detailed scores"),
    2021: ("2021-Government-AI-Readiness-Index-public-dataset.xlsx", "Detailed scores"),
    2022: ("2022-Government-AI-Readiness-Index-public-data.xlsx", "Detailed scores"),
    2023: ("2023-Government-AI-Readiness-Index-Public-Indicator-Data.xlsx", "Pillar & dimension scores"),
    2024: ("2024-GAIRI-data.xlsx", "Scores per pillar and dimension"),
}

FUENTES_2020_2024 = [
    ("total_score", "2020-2023", "B", 1, "Overall score", "Header: 'Overall score'. Promedio 3 pilares."),
    ("total_score", "2024", "B", 1, "Total", "Header: 'Total' (diferente a 2020-2023)."),
    ("government", "2020-2024", "C", 2, "Government", ""),
    ("technology_sector", "2020-2024", "D", 3, "Technology Sector", "Header con espacio."),
    ("data_infrastructure", "2020-2024", "E", 4, "Data and Infrastructure", ""),
    ("vision", "2020-2024", "F", 5, "Vision", "Gap entre Data&Infrastructure y Vision."),
    ("governance_ethics", "2020-2024", "G", 6, "Governance and Ethics", ""),
    ("digital_capacity", "2020-2024", "H", 7, "Digital Capacity", ""),
    ("adaptability", "2020-2024", "I", 8, "Adaptability", ""),
    ("size", "2020-2021", "J", 9, "Size", "SOLO 2020-2021. Renombrada a Maturity en 2022."),
    ("maturity", "2022-2024", "J", 9, "Maturity", "DESDE 2022. Antes (2020-2021) se llamaba Size."),
    ("innovation_capacity", "2020-2024", "K", 10, "Innovation Capacity", "2 missing en 2021."),
    ("human_capital", "2020-2024", "L", 11, "Human Capital", ""),
    ("infrastructure", "2020-2024", "M", 12, "Infrastructure", ""),
    ("data_availability", "2020-2024", "N", 13, "Data Availability", ""),
    ("data_representativeness", "2020-2024", "O", 14, "Data Representativeness",
     "Missing: 7(2020), 1(2021), 1(2022), 2(2023), 0(2024)."),
]

# ================================================================
# FUENTES_2025
# ================================================================
FUENTES_2025 = [
    ("rank", "A", 0, "Ranking", "Ordinal 1-195."),
    ("total_score", "C", 2, "Total Score", "Gap en col D."),
    ("policy_capacity", "E", 4, "Policy Capacity", ""),
    ("ai_infrastructure", "F", 5, "AI Infrastructure", ""),
    ("governance_2025", "G", 6, "Governance", "NO confundir con government de 2020-2024."),
    ("public_sector_adoption", "H", 7, "Public Sector Adoption", ""),
    ("development_diffusion", "I", 8, "Development & Diffusion", "Header incluye &."),
    ("resilience", "J", 9, "Resilience", ""),
    ("policy_vision", "L", 11, "Policy vision", "Gap en col K."),
    ("policy_commitment", "M", 12, "Policy commitment", ""),
    ("compute_capacity", "N", 13, "Compute capacity", ""),
    ("enabling_technical_infrastructure", "O", 14, "Enabling technical infrastructure", ""),
    ("data_quality", "P", 15, "Data quality", ""),
    ("governance_principles", "Q", 16, "Governance principles", ""),
    ("regulatory_compliance", "R", 17, "Regulatory compliance", ""),
    ("government_digital_policy", "S", 18, "Government digital policy", ""),
    ("e_government_delivery", "T", 19, "e-Government delivery", ""),
    ("human_capital_2025", "U", 20, "Human capital", "NO confundir con human_capital 2020-2024."),
    ("ai_sector_maturity", "V", 21, "AI sector maturity", ""),
    ("ai_technology_diffusion", "W", 22, "AI technology diffusion", ""),
    ("societal_transition", "X", 23, "Societal transition", ""),
    ("safety_security", "Y", 24, "Safety and security", ""),
]

# ================================================================
# FUENTES_2023_INDICATORS
# ================================================================
INDICATOR_2023_INFO = [
    ("ind_ai_strategy", 1, "B", "AI strategy", "desk research"),
    ("ind_data_protection_laws", 2, "C", "Data protection and privacy laws",
     "UN data protection and privacy legislation worldwide."),
    ("ind_cybersecurity", 3, "D", "Cybersecurity", "Global Cybersecurity Index"),
    ("ind_regulatory_quality", 4, "E", "Regulatory quality", "WGI Indicators"),
    ("ind_ethical_principles", 5, "F", "Ethical principles", "desk research"),
    ("ind_accountability", 6, "G", "Accountability", "WGI Indicators"),
    ("ind_online_services", 7, "H", "Online services", "UN E-gov Online Services Index"),
    ("ind_foundational_it", 8, "I", "Foundational IT infrastructure", "Govtech Maturity Index"),
    ("ind_govt_promotion_emerging_tech", 9, "J", "Government Promotion of Investment in Emerging Technologies", "Network Readiness Index"),
    ("ind_govt_effectiveness", 10, "K", "Government Effectiveness", "WGI Indicators"),
    ("ind_govt_responsiveness", 11, "L", "Government responsiveness to change", "Global Competitiveness Index"),
    ("ind_procurement_data", 12, "M", "Procurement Data", "Procurement page for the index"),
    ("ind_ai_unicorns_log", 13, "N", "Number of AI Unicorns log transformation", "CB Insights"),
    ("ind_non_ai_unicorns_log", 14, "O", "Number of non-AI Unicorns log transformation", "CB Insights"),
    ("ind_ict_trade_services_log", 15, "P", "Value of trade in ICT services (per capita) log transformation", "UNCTAD"),
    ("ind_ict_trade_goods_log", 16, "Q", "Value of trade in ICT goods (per capita) log transformation", "UNCTAD"),
    ("ind_software_spending", 17, "R", "Computer software spending", "Global Innovation Index"),
    ("ind_time_govt_regulations", 18, "S", "Time spent dealing with government regulations", "World Bank"),
    ("ind_vc_availability", 19, "T", "VC availability", "Global Innovation Index"),
    ("ind_rd_spending_log", 20, "U", "R&D Spending log transformation", "UNESCO"),
    ("ind_company_investment_emerging_tech", 21, "V", "Company investment in emerging technology", "Network Readiness Index"),
    ("ind_ai_research_papers_log", 22, "W", "AI research papers log transformation", "Scimago"),
    ("ind_stem_graduates", 23, "X", "Graduates in STEM or computer science", "UNESCO"),
    ("ind_github_activity_log", 24, "Y", "Github Activity log transformation", "Github 2021 Octoverse report"),
    ("ind_female_stem_graduates", 25, "Z", "Female STEM Graduates", "World Bank"),
    ("ind_eng_tech_higher_ed_quality", 26, "AA", "Quality of Engineering and Technology Higher Ed", "QS University Rankings"),
    ("ind_ict_skills", 27, "AB", "ICT skills", "ITU"),
    ("ind_telecom_infrastructure", 28, "AC", "Telecommunications Infrastructure", "UN E-gov Telecom Infra Index"),
    ("ind_supercomputers_log", 29, "AD", "Supercomputers log transformation", "Top 500 supercomputers"),
    ("ind_broadband_quality", 30, "AE", "Broadband Quality", "EIU Inclusive Internet Index"),
    ("ind_5g_infrastructure", 31, "AF", "5G Infrastructure", "Ookla 5G Map"),
    ("ind_adoption_emerging_tech", 32, "AG", "Adoption of Emerging Technologies", "Network Readiness Index"),
    ("ind_open_data", 33, "AH", "Open Data", "Global Data Barometer"),
    ("ind_data_governance", 34, "AI", "Data governance", "Govtech Maturity Index"),
    ("ind_mobile_subscriptions", 35, "AJ", "Mobile-cellular telephone subscriptions", "ITU"),
    ("ind_households_internet", 36, "AK", "Households with internet access", "ITU"),
    ("ind_statistical_capacity", 37, "AL", "Statistical Capacity", "SPI Github repo"),
    ("ind_cost_internet_device", 38, "AM", "Cost of cheapest internet-enabled device", "GSMA Mobile Connectivity Index"),
    ("ind_gender_gap_internet", 39, "AN", "Gender gap in internet access", "EIU Inclusive Internet Index"),
]

# ================================================================
# BUILD
# ================================================================
df_schema = pd.read_csv(output_path("schema_consolidado.csv"))

# --- Diccionario ---
dicc_rows = []
for _, srow in df_schema.iterrows():
    var, anos, fw, desc, tipo = srow["variable"], str(srow["anos"]), str(srow["framework"]), srow["descripcion"], srow["tipo"]
    if var in ["iso3","entity_type","pais_original","year","scale","framework","region_pertenece"]:
        escala = "string"
    elif var == "privacy_laws":
        escala = "binario (0/1)"
    elif var == "ai_strategy" and not var.startswith("ind_"):
        escala = "ordinal (0/1/2)"
    elif var == "ai_startups":
        escala = "conteo (0-5053)"
    elif var.startswith("norm_") or var.startswith("avg_"):
        escala = "0-1"
    elif var.startswith("ind_"):
        escala = "0-100"
    elif var == "total_score" and "2019" in anos and "2020" not in anos:
        escala = "0-10"
    elif var in ["rank","ranking_detail"] or var.startswith("rank_"):
        escala = "ordinal (1=N)"
    elif var.startswith("score_"):
        escala = "0-10"
    else:
        escala = "0-100"
    dicc_rows.append({"variable":var,"hoja_destino":"Consolidado + hojas de detalle","descripcion":desc,
                      "entidad":"pais","tipo_valor":tipo,"escala":escala,
                      "anos_disponibles":anos,"framework":fw,"notas":""})

df_dicc = pd.DataFrame(dicc_rows)

# --- Fuentes ---
fuente_rows = []
styles = {}

# Entity columns
for var, arch, hoja, col, head, filas, notas in [
    ("iso3","UNIFICADO/src/iso3_mapping.py","iso3_mapping.csv","Columna iso3",
     "Mapeo manual desde nombres en 7 Excels","229 entradas","DERIVADA."),
    ("entity_type","UNIFICADO/src/iso3_mapping.py","iso3_mapping.csv","Columna entity_type",
     "Derivado del mapeo manual","229 entradas","DERIVADA."),
    ("pais_original","TODOS los 7 Excels",
     "Data / Detailed scores / Pillar & dim / Scores per pillar / Dimensions-Pillars",
     "Col A (Country)","Nombre exacto en el Excel de cada ano","194-195/ano","No normalizado."),
    ("year","N/A (metadato)","N/A","N/A","Ano de edicion","N/A","DERIVADA."),
    ("scale","N/A (metadato)","N/A","N/A","Escala: 0-10(2019) o 0-100(2020-2025)","N/A","DERIVADA."),
    ("framework","N/A (metadato)","N/A","N/A","Framework del ano","N/A","DERIVADA."),
]:
    fuente_rows.append({"variable":var,"archivo_origen":arch,"hoja_origen":hoja,
                        "columna_origen":col,"header_origen_exacto":head,
                        "filas_datos":filas,"notas_lectura":notas})

# 2019 total_score
fuente_rows.append({"variable":"total_score","archivo_origen":"SHARED_-2019-Index-data-for-report.xlsx",
    "hoja_origen":"Data","columna_origen":"Col AA (indice 26)",
    "header_origen_exacto":"INDEX SCORE",
    "filas_datos":"Filas 7-200 (194 paises)",
    "notas_lectura":"Escala 0-10. NO es 0-100 como 2020-2025. Se renombro a total_score en Consolidado."})

# 2019 Data
for var, cl, ci, he, fd, nt in FUENTES_2019_DATA:
    fuente_rows.append({"variable":var,"archivo_origen":"SHARED_-2019-Index-data-for-report.xlsx",
        "hoja_origen":"Data",
        "columna_origen":f"Col {cl} (indice {ci})",
        "header_origen_exacto":he,
        "filas_datos":f"Filas {fd} (194 paises)",
        "notas_lectura":f"Fila 1=header. Fila 5=Max Score. {nt}"})

# 2019 Rankings
for var, cl, ci, he, fd, nt in FUENTES_2019_RANKINGS:
    cs = f"Col {cl} (indice {ci})" if cl != "derivado" else "Variable derivada"
    fuente_rows.append({"variable":var,"archivo_origen":"SHARED_-2019-Index-data-for-report.xlsx",
        "hoja_origen":"Rankings","columna_origen":cs,
        "header_origen_exacto":he,"filas_datos":f"Filas {fd} (194 paises)",
        "notas_lectura":f"Fila 0=regiones, Fila 1=sub-headers. {nt}"})

# 2020-2024
for v, ys, cl, ci, he, nt in FUENTES_2020_2024:
    if "2020-2024" in ys: yr_list = [2020,2021,2022,2023,2024]
    elif "2020-2023" in ys: yr_list = [2020,2021,2022,2023]
    elif "2022-2024" in ys: yr_list = [2022,2023,2024]
    elif "2020-2021" in ys: yr_list = [2020,2021]
    elif ys=="2024": yr_list = [2024]
    else: continue
    for yr in yr_list:
        yf, ys2 = YEAR_SHEET_MAP[yr]
        fuente_rows.append({"variable":v,"archivo_origen":yf,"hoja_origen":ys2,
            "columna_origen":f"Col {cl} (indice {ci} tras limpiar unnamed)",
            "header_origen_exacto":he,
            "filas_datos":f"Filas 3-{EXPECTED_COUNTS[yr]+2} ({EXPECTED_COUNTS[yr]} paises)",
            "notas_lectura":f"Header fila 2. Unnamed eliminadas. {nt}"})

# 2024 ranking_detail
fuente_rows.append({"variable":"ranking_detail","archivo_origen":"2024-GAIRI-data.xlsx",
    "hoja_origen":"Scores per pillar and dimension",
    "columna_origen":"Col C (indice 2)","header_origen_exacto":"Ranking",
    "filas_datos":"Filas 3-190 (188 paises, 82 sin ranking)",
    "notas_lectura":"SOLO 2024. 82 missing (106 rankeados)."})

# 2025
for v, cl, ci, he, nt in FUENTES_2025:
    fuente_rows.append({"variable":v,"archivo_origen":"2025-Government-AI-Readiness-Index-data-1.xlsx",
        "hoja_origen":"Dimensions-Pillars",
        "columna_origen":f"Col {cl} (indice {ci} tras eliminar gaps)",
        "header_origen_exacto":he,
        "filas_datos":"Filas 3-197 (195 paises)",
        "notas_lectura":f"Fila 1=labels, Fila 2=header REAL. {nt}"})

# 2023 Indicators
for v, ci, cl, ind_name, ind_src in INDICATOR_2023_INFO:
    fuente_rows.append({"variable":v,"archivo_origen":"2023-Government-AI-Readiness-Index-Public-Indicator-Data.xlsx",
        "hoja_origen":"Indicator scores",
        "columna_origen":f"Col {cl} (indice {ci})",
        "header_origen_exacto":ind_name,
        "filas_datos":"Filas 5-197 (193 paises)",
        "notas_lectura":f"Fila 3=Indicator names. Fila 4=Source. Fuente original: {ind_src}"})

df_fuentes = pd.DataFrame(fuente_rows)

# Save CSVs
df_dicc.to_csv(output_path("diccionario_variables.csv"), index=False)
df_fuentes.to_csv(output_path("fuentes_directas.csv"), index=False)
print(f"Diccionario: {len(df_dicc)} variables")
print(f"Fuentes_Directas v2: {len(df_fuentes)} filas")

# ================================================================
# UPDATE EXCEL: solo hojas Fuentes_Directas y Diccionario_Variables
# ================================================================
xlsx_path = os.path.join(os.path.dirname(OUTPUT_DIR), "Oxford_Insights_Unificado.xlsx")
print(f"\nActualizando solo las hojas de documentacion en: {xlsx_path}")

wb = load_workbook(xlsx_path)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align = Alignment(vertical="top", wrap_text=False)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def write_sheet(ws, df):
    """Write a dataframe to a worksheet, replacing existing content."""
    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border
    
    for ri, (_, row) in enumerate(df.iterrows()):
        for ci, h in enumerate(headers):
            v = row[h]
            c = ws.cell(row=ri+2, column=ci+1, value=None if pd.isna(v) else v)
            c.alignment = cell_align; c.border = thin_border
    
    for ci, h in enumerate(headers, 1):
        max_len = max(len(str(h)), df[h].astype(str).str.len().max() if len(df)>0 else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df)+1}"

# Update Fuentes_Directas
ws_fuentes = wb["Fuentes_Directas"]
write_sheet(ws_fuentes, df_fuentes)
print(f"  Fuentes_Directas actualizada: {len(df_fuentes)} filas x {len(df_fuentes.columns)} cols")

# Update Diccionario_Variables
ws_dicc = wb["Diccionario_Variables"]
write_sheet(ws_dicc, df_dicc)
print(f"  Diccionario_Variables actualizada: {len(df_dicc)} filas x {len(df_dicc.columns)} cols")

wb.save(xlsx_path)
print(f"\nWorkbook actualizado: {xlsx_path}")
print("NOTA: Solo se modificaron las hojas Fuentes_Directas y Diccionario_Variables.")
print("Las hojas de datos (Consolidado, Detalle_*, etc.) NO fueron tocadas.")
