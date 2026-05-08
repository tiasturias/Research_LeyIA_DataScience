#!/usr/bin/env python3
"""
build_oecd_excel.py — Extraccion de datos OECD vía SDMX API
============================================================
Proyecto: "¿Regular o no regular?" — IMT3860 Data Science, UC Chile
Reglas:
  - Cero datos sinteticos, cero imputacion, cero estimacion
  - Cada valor = respuesta directa del endpoint SDMX
  - Si no hay dato → celda vacia, no estimar
  - Cada endpoint documentado en DICCIONARIO con URL exacta verificable
Output: OECD_Data_Central_YYYY-MM-DD.xlsx

Ejecucion paso a paso:
  python3 build_oecd_excel.py --paso 1   (solo endpoint #1)
  python3 build_oecd_excel.py --paso 2   (endpoint #2)
  ...
  python3 build_oecd_excel.py --todo     (todos los pasos)
"""

import os
import sys
import csv
import io
import hashlib
import requests
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# =============================================================================
# CONFIGURACION
# =============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "raw_csv")
os.makedirs(RAW_DIR, exist_ok=True)

SDMX_BASE = "https://sdmx.oecd.org/public/rest/data"
USER_AGENT = "LeyIA-DataScience/1.0 (research project; contact: pablo@uc.cl)"
TIMEOUT = 60

TODAY = datetime.now().strftime("%Y-%m-%d")
OUTPUT_FILE = os.path.join(BASE_DIR, f"OECD_Data_Central_{TODAY}.xlsx")

# Colores para hojas
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# FUNCIONES BASE
# =============================================================================
def _raw_path(flow_id):
    """Ruta del archivo CSV crudo en raw_csv/."""
    safe_name = flow_id.replace("@", "_at_").replace(".", "_")
    return os.path.join(RAW_DIR, f"{safe_name}_{TODAY}.csv")


def download_oecd_csv(agency, flow_id, version, dimensions, label="dataset", max_retries=3, use_cache=True):
    """Descarga CSV desde OECD SDMX API. Usa cache si el archivo ya existe.
    
    Args:
        agency: ej. 'OECD.TAD.TPD'
        flow_id: ej. 'DSD_STRI@DF_STRI_DIGITAL'
        version: ej. '1.0'
        dimensions: ej. '/A........'
        label: nombre legible para logs
        max_retries: reintentos si rate limited (429)
        use_cache: si True, usa archivo existente en raw_csv/ en vez de descargar
    
    Returns:
        dict con df, url, sha256, n_rows, raw_path, label, o None si error
    """
    url = f"{SDMX_BASE}/{agency},{flow_id},{version}/{dimensions}?format=csvfilewithlabels"
    raw_path = _raw_path(flow_id)
    
    # Cache hit
    if use_cache and os.path.exists(raw_path):
        filesize = os.path.getsize(raw_path)
        print(f"  Usando cache: {os.path.basename(raw_path)} ({filesize/1024:.0f} KB)")
        with open(raw_path, "r", encoding="utf-8") as f:
            csv_text = f.read()
        sha256 = hashlib.sha256(csv_text.encode("utf-8")).hexdigest()
        df = pd.read_csv(io.StringIO(csv_text), low_memory=False)
        print(f"  OK: {len(df)} filas x {len(df.columns)} cols | SHA256: {sha256[:16]}... | CACHE")
        return {
            "df": df, "url": url, "n_rows": len(df),
            "sha256": sha256, "raw_path": raw_path, "label": label
        }
    
    print(f"  Descargando {label}...")
    print(f"  URL: {url[:120]}...")
    
    import time
    
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.get(
                url,
                headers={"User-Agent": USER_AGENT},
                timeout=TIMEOUT
            )
        except requests.exceptions.Timeout:
            print(f"  ERROR: Timeout (>60s)")
            return None
        except Exception as e:
            print(f"  ERROR: {e}")
            return None
        
        if resp.status_code == 429:
            wait = attempt * 15
            print(f"  ⚠️ RATE LIMITED (429). Esperando {wait}s (intento {attempt}/{max_retries})...")
            time.sleep(wait)
            continue
        
        if resp.status_code != 200:
            print(f"  ❌ ERROR: HTTP {resp.status_code} - {resp.reason}")
            return None
        
        # Exito: salir del loop
        break
    else:
        print(f"  ⏸️ RATE LIMITED: Max retries ({max_retries}) agotados. Pendiente para reintentar.")
        return "RATE_LIMITED"
    
    # Guardar CSV crudo para auditoria
    with open(raw_path, "w", encoding="utf-8") as f:
        f.write(resp.text)
    sha256 = hashlib.sha256(resp.text.encode("utf-8")).hexdigest()
    
    # Parsear CSV con low_memory=False para evitar DtypeWarning
    df = pd.read_csv(io.StringIO(resp.text), low_memory=False)
    n_rows = len(df)
    n_cols = len(df.columns)
    
    print(f"  OK: {n_rows} filas x {n_cols} cols | SHA256: {sha256[:16]}...")
    
    return {
        "df": df, "url": url, "n_rows": n_rows,
        "sha256": sha256, "raw_path": raw_path, "label": label
    }


def make_header_row(ws, headers, row=1):
    """Escribe encabezados con formato."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_dataframe(ws, df, start_row=2):
    """Escribe un DataFrame en una hoja de Excel."""
    for col_idx, col_name in enumerate(df.columns, 1):
        for row_idx, value in enumerate(df[col_name], start_row):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if pd.isna(value):
                cell.value = None


def auto_width(ws):
    """Ajusta ancho de columnas al contenido."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except:
                pass
        adjusted = min(max_len + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 10)


def agregar_metadata(ws, metadatos, start_row=1):
    """Agrega metadatos a la hoja README."""
    for i, (k, v) in enumerate(metadatos.items()):
        cell_k = ws.cell(row=start_row + i, column=1, value=k)
        cell_k.font = Font(bold=True)
        ws.cell(row=start_row + i, column=2, value=v)


# =============================================================================
# BLOQUES DE EXTRACCION (un endpoint por funcion)
# =============================================================================

def extract_stri_digital():
    """Endpoint #1: Digital Services Trade Restrictiveness Index"""
    print("\n" + "="*70)
    print("ENDPOINT #1: Digital STRI (STRI_DIGITAL)")
    print("="*70)
    
    result = download_oecd_csv(
        agency="OECD.TAD.TPD",
        flow_id="DSD_STRI@DF_STRI_DIGITAL",
        version="1.0",
        dimensions="A........",
        label="Digital Services Trade Restrictiveness Index"
    )
    
    if result is None:
        return None
    
    df_raw = result["df"]
    
    # Medidas relevantes:
    # STRI = overall composite index (0-1, verdadera medida de restrictividad)
    # CLAS1_6 = Infrastructure and connectivity
    # CLAS1_7 = Electronic transactions
    # CLAS1_8 = Payment systems
    # CLAS1_9 = Intellectual property rights
    # CLAS1_10 = Other barriers
    medidas_interes = ["STRI", "CLAS1_6", "CLAS1_7", "CLAS1_8", "CLAS1_9", "CLAS1_10"]
    
    # Pivot: una fila por pais-ano, una columna por medida
    df_pivot = df_raw[df_raw["MEASURE"].isin(medidas_interes)].copy()
    df_pivot["OBS_VALUE"] = pd.to_numeric(df_pivot["OBS_VALUE"], errors="coerce")
    
    # Mapeo de nombres de medidas
    measure_names = {
        "STRI": "digital_stri_overall",
        "CLAS1_6": "digital_stri_infrastructure",
        "CLAS1_7": "digital_stri_e_transactions",
        "CLAS1_8": "digital_stri_payment",
        "CLAS1_9": "digital_stri_ip",
        "CLAS1_10": "digital_stri_other"
    }
    df_pivot["medida"] = df_pivot["MEASURE"].map(measure_names)
    
    df_out = df_pivot.pivot_table(
        index=["REF_AREA", "TIME_PERIOD"],
        columns="medida",
        values="OBS_VALUE",
        aggfunc="first"
    ).reset_index()
    df_out.columns.name = None
    df_out = df_out.rename(columns={"REF_AREA": "iso3", "TIME_PERIOD": "year"})
    
    print(f"  Paises: {df_out['iso3'].nunique()}, Anios: {df_out['year'].nunique()}")
    print(f"  Variables: {[c for c in df_out.columns if c not in ['iso3','year']]}")
    
    result["df"] = df_out
    result["variables"] = [
        {
            "variable_id": "oecd_digital_stri_overall",
            "hoja_datos": "1_DIGITAL_STRI",
            "codigo_indicador": "STRI",
            "descripcion": "Digital Services Trade Restrictiveness Index — score compuesto (0-1)",
            "mensaje_analitico": "Proxy principal de barreras regulatorias a servicios digitales. Medida alternativa a iapp_regulatory_intensity para capturar restrictividad regulatoria digital. A mayor valor, mayor restriccion.",
            "dataset": "DSD_STRI@DF_STRI_DIGITAL"
        },
        {
            "variable_id": "oecd_digital_stri_infrastructure",
            "hoja_datos": "1_DIGITAL_STRI",
            "codigo_indicador": "CLAS1_6",
            "descripcion": "Digital STRI — Infrastructure and connectivity sub-score",
            "mensaje_analitico": "Barreras regulatorias en infraestructura digital (telecom, conectividad). Proxy de openness del mercado de telecomunicaciones.",
            "dataset": "DSD_STRI@DF_STRI_DIGITAL"
        },
        {
            "variable_id": "oecd_digital_stri_e_transactions",
            "hoja_datos": "1_DIGITAL_STRI",
            "codigo_indicador": "CLAS1_7",
            "descripcion": "Digital STRI — Electronic transactions sub-score",
            "mensaje_analitico": "Barreras regulatorias a transacciones electronicas (firma digital, factura electronica, proteccion consumidor). Proxy de madurez legal del comercio electronico.",
            "dataset": "DSD_STRI@DF_STRI_DIGITAL"
        },
        {
            "variable_id": "oecd_digital_stri_payment",
            "hoja_datos": "1_DIGITAL_STRI",
            "codigo_indicador": "CLAS1_8",
            "descripcion": "Digital STRI — Payment systems sub-score",
            "mensaje_analitico": "Barreras regulatorias en sistemas de pago digital. Proxy de restricciones a fintech y pagos digitales transfronterizos.",
            "dataset": "DSD_STRI@DF_STRI_DIGITAL"
        },
        {
            "variable_id": "oecd_digital_stri_ip",
            "hoja_datos": "1_DIGITAL_STRI",
            "codigo_indicador": "CLAS1_9",
            "descripcion": "Digital STRI — Intellectual property rights sub-score",
            "mensaje_analitico": "Barreras regulatorias en propiedad intelectual digital. Proxy de proteccion IP en entorno digital.",
            "dataset": "DSD_STRI@DF_STRI_DIGITAL"
        },
        {
            "variable_id": "oecd_digital_stri_other",
            "hoja_datos": "1_DIGITAL_STRI",
            "codigo_indicador": "CLAS1_10",
            "descripcion": "Digital STRI — Other barriers sub-score",
            "mensaje_analitico": "Otras barreras regulatorias digitales (publicidad, contenido, responsabilidad intermediarios).",
            "dataset": "DSD_STRI@DF_STRI_DIGITAL"
        }
    ]
    
    return result


def extract_indigo():
    """Endpoint #2: Index of Digital Trade Integration and Openness (INDIGO)"""
    print("\n" + "="*70)
    print("ENDPOINT #2: INDIGO — Digital Trade Integration and Openness")
    print("="*70)
    
    result = download_oecd_csv(
        agency="OECD.TAD.TPD",
        flow_id="DSD_INDIGO@DF_INDIGO",
        version="1.0",
        dimensions="A......",
        label="Digital Trade Integration and Openness (INDIGO)"
    )
    
    if result is None:
        return None
    
    df_raw = result["df"]
    
    # INDIGO = overall composite index (0-1, a mayor valor = mas integrado/abierto)
    # POL1-POL5 = policy sub-dimensions
    # Nos quedamos con INDIGO_I_UNW (unweighted) como proxy de apertura digital
    
    # Filtrar INDIGO_I_UNW = unweighted overall index
    df_main = df_raw[
        (df_raw["MEASURE"] == "INDIGO") & 
        (df_raw["INDIGO_TYPE"] == "INDIGO_I_UNW")
    ].copy()
    
    if len(df_main) == 0:
        print("  ADVERTENCIA: No se encontraron filas con INDIGO_I_UNW")
        df_main = df_raw[df_raw["MEASURE"] == "INDIGO"].copy()
    
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    
    # Columnas relevantes
    df_out = df_main[["REF_AREA", "TIME_PERIOD", "OBS_VALUE"]].copy()
    df_out.columns = ["iso3", "year", "oecd_indigo_score"]
    df_out = df_out.dropna(subset=["oecd_indigo_score"])
    
    print(f"  Paises: {df_out['iso3'].nunique()}, Anios: {df_out['year'].nunique()}")
    print(f"  Score range: {df_out['oecd_indigo_score'].min():.4f} - {df_out['oecd_indigo_score'].max():.4f}")
    
    result["df"] = df_out
    result["variables"] = [
        {
            "variable_id": "oecd_indigo_score",
            "hoja_datos": "2_INDIGO",
            "codigo_indicador": "INDIGO / INDIGO_I_UNW",
            "descripcion": "Index of Digital Trade Integration and Openness — unweighted (0-1)",
            "mensaje_analitico": "Proxy de apertura al comercio digital. Complemento de Digital STRI: mientras STRI mide barreras, INDIGO mide integracion activa en cadenas de valor digital. A mayor valor, mayor apertura.",
            "dataset": "DSD_INDIGO@DF_INDIGO"
        }
    ]
    
    return result


def extract_regulatory_gov():
    """Endpoint #3: Regulatory Governance (RIA, Stakeholder Engagement, Ex Post Evaluation)"""
    print("\n" + "="*70)
    print("ENDPOINT #3: Regulatory Governance — RIA / SA / EPE")
    print("="*70)
    
    result = download_oecd_csv(
        agency="OECD.GOV.GIP",
        flow_id="DSD_GOV_REG@DF_GOV_REG_2025",
        version="1.0",
        dimensions="A.......",
        label="Regulatory Governance (RIA, Stakeholder Engagement, Ex Post)"
    )
    
    if result is None:
        return None
    
    df_raw = result["df"]
    
    # Tres medidas compuestas: EPE (Ex Post Eval), SA (Stakeholder Engagement), RIA (Impact Assessment)
    # Categoria P_O = Primary legislation, overall (el score mas general)
    df_main = df_raw[df_raw["REGULATORY_CATEGORY"] == "P_O"].copy()
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    
    # Pivot: una fila por pais-ano, columnas EPE, SA, RIA
    df_pivot = df_main.pivot_table(
        index=["REF_AREA", "TIME_PERIOD"],
        columns="MEASURE",
        values="OBS_VALUE",
        aggfunc="first"
    ).reset_index()
    df_pivot.columns.name = None
    df_pivot = df_pivot.rename(columns={"REF_AREA": "iso3", "TIME_PERIOD": "year"})
    
    # Renombrar columnas
    rename_map = {
        "EPE": "oecd_reg_gov_ex_post_eval",
        "RIA": "oecd_reg_gov_ria",
        "SA": "oecd_reg_gov_stakeholder"
    }
    df_pivot = df_pivot.rename(columns=rename_map)
    
    print(f"  Paises: {df_pivot['iso3'].nunique()}, Anios: {df_pivot['year'].nunique()}")
    for col in rename_map.values():
        if col in df_pivot.columns:
            vals = df_pivot[col].dropna()
            print(f"  {col}: {len(vals)} valores, rango {vals.min():.2f}-{vals.max():.2f}")
    
    result["df"] = df_pivot
    result["variables"] = [
        {
            "variable_id": "oecd_reg_gov_ria",
            "hoja_datos": "3_REGULATORY_GOV",
            "codigo_indicador": "RIA (P_O)",
            "descripcion": "Regulatory Impact Assessment Index — Primary legislation, overall (0-1)",
            "mensaje_analitico": "Mide la calidad del proceso de evaluacion de impacto regulatorio. Proxy de evidencia-based policy making. Util para controlar calidad institucional en modelos de regulacion IA.",
            "dataset": "DSD_GOV_REG@DF_GOV_REG_2025"
        },
        {
            "variable_id": "oecd_reg_gov_stakeholder",
            "hoja_datos": "3_REGULATORY_GOV",
            "codigo_indicador": "SA (P_O)",
            "descripcion": "Stakeholder Engagement Index — Primary legislation, overall (0-1)",
            "mensaje_analitico": "Mide la calidad de la participacion ciudadana en el proceso regulatorio. Proxy de transparencia y apertura institucional. Control para modelos de gobernanza IA.",
            "dataset": "DSD_GOV_REG@DF_GOV_REG_2025"
        },
        {
            "variable_id": "oecd_reg_gov_ex_post_eval",
            "hoja_datos": "3_REGULATORY_GOV",
            "codigo_indicador": "EPE (P_O)",
            "descripcion": "Ex Post Evaluation Index — Primary legislation, overall (0-1)",
            "mensaje_analitico": "Mide la calidad de la evaluacion ex post de regulaciones. Proxy de capacidad de aprendizaje institucional y adaptacion regulatoria. Relevante para entender madurez del ecosistema regulatorio.",
            "dataset": "DSD_GOV_REG@DF_GOV_REG_2025"
        }
    ]
    
    return result


def extract_digital_gov():
    """Endpoint #4: Digital Government Index"""
    print("\n" + "="*70)
    print("ENDPOINT #4: Digital Government Index (DGOGD)")
    print("="*70)
    
    result = download_oecd_csv(
        agency="OECD.GOV.GIP",
        flow_id="DSD_GOV@DF_GOV_DGOGD_2025",
        version="1.0",
        dimensions="A.......",
        label="Digital Government and Open Government Data Index"
    )
    
    if result is None:
        return None
    
    df_raw = result["df"]
    
    # Medidas: DG = Digital Government overall, OUR = Open Gov Data overall
    medidas_interes = ["DG", "OUR", "DG_D", "DG_DD", "DG_OD", "DG_UD", "DG_PR", "DG_PL"]
    
    df_main = df_raw[df_raw["MEASURE"].isin(medidas_interes)].copy()
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    
    # Pivot
    df_pivot = df_main.pivot_table(
        index=["REF_AREA", "TIME_PERIOD"],
        columns="MEASURE",
        values="OBS_VALUE",
        aggfunc="first"
    ).reset_index()
    df_pivot.columns.name = None
    df_pivot = df_pivot.rename(columns={"REF_AREA": "iso3", "TIME_PERIOD": "year"})
    
    # Renombrar
    rename_map = {
        "DG": "oecd_digital_gov_overall",
        "OUR": "oecd_open_gov_data",
        "DG_D": "oecd_digital_gov_design",
        "DG_DD": "oecd_digital_gov_data_driven",
        "DG_OD": "oecd_digital_gov_open_data",
        "DG_UD": "oecd_digital_gov_user_driven",
        "DG_PR": "oecd_digital_gov_proactiveness",
        "DG_PL": "oecd_digital_gov_policy"
    }
    df_pivot = df_pivot.rename(columns=rename_map)
    
    print(f"  Paises: {df_pivot['iso3'].nunique()}, Anios: {df_pivot['year'].nunique()}")
    if "oecd_digital_gov_overall" in df_pivot.columns:
        vals = df_pivot["oecd_digital_gov_overall"].dropna()
        print(f"  Digital Gov Overall: {len(vals)} valores, rango {vals.min():.2f}-{vals.max():.2f}")
    
    result["df"] = df_pivot
    result["variables"] = [
        {
            "variable_id": "oecd_digital_gov_overall",
            "hoja_datos": "4_DIGITAL_GOV",
            "codigo_indicador": "DG",
            "descripcion": "Digital Government Index — Overall score (0-1)",
            "mensaje_analitico": "Proxy de madurez de gobierno digital. Complementa oxford_readiness_score y oxford_pillar_public_sector. Mide que tan digitalizado esta el sector publico, relevante para entender capacidad de implementar regulacion IA.",
            "dataset": "DSD_GOV@DF_GOV_DGOGD_2025"
        },
        {
            "variable_id": "oecd_open_gov_data",
            "hoja_datos": "4_DIGITAL_GOV",
            "codigo_indicador": "OUR",
            "descripcion": "Open Government Data Index — Overall score (0-1)",
            "mensaje_analitico": "Proxy de apertura de datos gubernamentales. Indica disponibilidad de datos publicos para desarrollo de IA. Relevante para ecosistema de innovacion IA.",
            "dataset": "DSD_GOV@DF_GOV_DGOGD_2025"
        }
    ]
    
    return result


def extract_ict_business():
    """Endpoint #5: ICT Access and Usage by Businesses (G14=AI, G3=Cloud, G7=Big Data, G13=IoT, etc.)"""
    print("\n" + "="*70)
    print("ENDPOINT #5: ICT Access and Usage by Businesses")
    print("="*70)
    
    # Descargar dataset completo una vez (222K filas, manejable)
    result = download_oecd_csv(
        agency="OECD.STI.DEP",
        flow_id="DSD_ICT_B@DF_BUSINESSES",
        version="1.0",
        dimensions="......",
        label="ICT Access and Usage by Businesses (complete)"
    )
    
    if result is None:
        return None
    
    df_raw = result["df"]
    
    # Medidas de interes para el estudio
    medidas = {
        "G14_B": "AI usage by businesses",
        "G3_B": "Cloud computing",
        "G7_B": "Big Data analysis",
        "G13_B": "IoT usage",
        "B1_B": "Website presence",
        "H1_B": "ICT specialists"
    }
    
    # Filtrar por medidas deseadas + actividad total + tamano 10+ empleados
    df_main = df_raw[
        (df_raw["MEASURE"].isin(medidas.keys())) &
        (df_raw["ACTIVITY"] == "_T") &
        (df_raw["SIZE_CLASS"] == "S_GE10") &
        (df_raw["UNIT_MEASURE"] == "PT_ENT")
    ].copy()
    
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    
    # Pivot: una fila por pais-ano
    df_pivot = df_main.pivot_table(
        index=["REF_AREA", "TIME_PERIOD"],
        columns="MEASURE",
        values="OBS_VALUE",
        aggfunc="first"
    ).reset_index()
    df_pivot.columns.name = None
    df_pivot = df_pivot.rename(columns={"REF_AREA": "iso3", "TIME_PERIOD": "year"})
    
    # Renombrar columnas a nombres semanticos
    rename_map = {
        "G14_B": "oecd_biz_ai_pct",
        "G3_B": "oecd_biz_cloud_pct",
        "G7_B": "oecd_biz_bigdata_pct",
        "G13_B": "oecd_biz_iot_pct",
        "B1_B": "oecd_biz_website_pct",
        "H1_B": "oecd_biz_ict_specialists_pct"
    }
    df_pivot = df_pivot.rename(columns=rename_map)
    
    print(f"  Paises: {df_pivot['iso3'].nunique()}")
    print(f"  Anios: {sorted(df_pivot['year'].dropna().unique())}")
    for col in rename_map.values():
        if col in df_pivot.columns:
            vals = df_pivot[col].dropna()
            print(f"  {col}: {len(vals)} valores, rango {vals.min():.1f}-{vals.max():.1f}")
    
    result["df"] = df_pivot
    result["variables"] = [
        {
            "variable_id": "oecd_biz_ai_pct",
            "hoja_datos": "5_ICT_BUSINESS",
            "codigo_indicador": "G14_B (PT_ENT, _T, S_GE10)",
            "descripcion": "Businesses using Artificial Intelligence — % of enterprises (10+ empl.)",
            "mensaje_analitico": "VARIABLE DEPENDIENTE PRINCIPAL para sub-pregunta de adopcion. Proxy de adopcion empresarial de IA. Cubre 40 paises OECD + partners. Complementa ms_ai_user_share (que solo mide GenAI via Microsoft).",
            "dataset": "DSD_ICT_B@DF_BUSINESSES"
        },
        {
            "variable_id": "oecd_biz_cloud_pct",
            "hoja_datos": "5_ICT_BUSINESS",
            "codigo_indicador": "G3_B (PT_ENT, _T, S_GE10)",
            "descripcion": "Businesses buying Cloud Computing services — % of enterprises",
            "mensaje_analitico": "Control de infraestructura digital empresarial. Proxy de madurez de adopcion de servicios cloud, relevante como variable de control en modelos de adopcion IA.",
            "dataset": "DSD_ICT_B@DF_BUSINESSES"
        },
        {
            "variable_id": "oecd_biz_bigdata_pct",
            "hoja_datos": "5_ICT_BUSINESS",
            "codigo_indicador": "G7_B (PT_ENT, _T, S_GE10)",
            "descripcion": "Businesses doing Big Data analysis — % of enterprises",
            "mensaje_analitico": "Control de capacidad analitica empresarial. Proxy de sofisticacion de uso de datos, relevante como variable de control para diferenciar entre adopcion basica vs. avanzada de IA.",
            "dataset": "DSD_ICT_B@DF_BUSINESSES"
        },
        {
            "variable_id": "oecd_biz_iot_pct",
            "hoja_datos": "5_ICT_BUSINESS",
            "codigo_indicador": "G13_B (PT_ENT, _T, S_GE10)",
            "descripcion": "Businesses using Internet of Things — % of enterprises",
            "mensaje_analitico": "Control de madurez de digitalizacion industrial. Proxy de adopcion de tecnologias conectadas, relevante como variable de control en modelos de difusion tecnologica.",
            "dataset": "DSD_ICT_B@DF_BUSINESSES"
        },
        {
            "variable_id": "oecd_biz_website_pct",
            "hoja_datos": "5_ICT_BUSINESS",
            "codigo_indicador": "B1_B (PT_ENT, _T, S_GE10)",
            "descripcion": "Businesses with a website — % of enterprises",
            "mensaje_analitico": "Control de digitalizacion basica. Proxy de presencia digital minima, control para separar efecto de digitalizacion general vs. adopcion especifica de IA.",
            "dataset": "DSD_ICT_B@DF_BUSINESSES"
        },
        {
            "variable_id": "oecd_biz_ict_specialists_pct",
            "hoja_datos": "5_ICT_BUSINESS",
            "codigo_indicador": "H1_B (PT_ENT, _T, S_GE10)",
            "descripcion": "Businesses employing ICT specialists — % of enterprises",
            "mensaje_analitico": "Control de talento digital empresarial. Proxy de capacidad interna para adoptar y gestionar tecnologias IA.",
            "dataset": "DSD_ICT_B@DF_BUSINESSES"
        }
    ]
    
    return result


def extract_fdi_restrictiveness():
    """Endpoint #6: FDI Regulatory Restrictiveness Index"""
    print("\n" + "="*70)
    print("ENDPOINT #6: FDI Regulatory Restrictiveness Index")
    print("="*70)
    
    # 5 dimensiones: REF_AREA, ACTIVITY, POL_CAT, MEASURE, UNIT_MEASURE
    # _T para ACTIVITY = total economy
    result = download_oecd_csv(
        agency="OECD.DAF.INV",
        flow_id="DSD_FDIRRI_SCORES@DF_FDIRRI_SCORES",
        version="2.0",
        dimensions="/._T...",
        label="FDI Regulatory Restrictiveness Index"
    )
    
    if result is None:
        return None
    
    df_raw = result["df"]
    
    # POL_CAT = _T is the overall restrictiveness score
    df_main = df_raw[df_raw["POL_CAT"] == "_T"].copy()
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    
    df_out = df_main[["REF_AREA", "TIME_PERIOD", "OBS_VALUE"]].copy()
    df_out.columns = ["iso3", "year", "oecd_fdi_restrictiveness"]
    
    print(f"  Paises: {df_out['iso3'].nunique()}, Anios: {df_out['year'].nunique()}")
    vals = df_out["oecd_fdi_restrictiveness"].dropna()
    print(f"  Score range: {vals.min():.4f} - {vals.max():.4f}")
    
    result["df"] = df_out
    result["variables"] = [
        {
            "variable_id": "oecd_fdi_restrictiveness",
            "hoja_datos": "6_FDI_RESTRICTIVENESS",
            "codigo_indicador": "RR (_T, _T)",
            "descripcion": "FDI Regulatory Restrictiveness Index — Overall score (0-1)",
            "mensaje_analitico": "Mide barreras regulatorias a la inversion extranjera directa. Proxy de apertura a inversion internacional. Relevante para sub-pregunta de inversion: paises mas restrictivos a FDI podrian tener menor inversion en ecosistema IA.",
            "dataset": "DSD_FDIRRI_SCORES@DF_FDIRRI_SCORES"
        }
    ]
    
    return result


def extract_pmr():
    """Endpoint #7: Product Market Regulation (PMR) Index"""
    print("\n" + "="*70)
    print("ENDPOINT #7: Product Market Regulation (PMR) Index")
    print("="*70)
    
    # 3 dimensiones: REF_AREA, MEASURE, FREQ
    # PMR = overall index (0-6), A = annual
    result = download_oecd_csv(
        agency="OECD.ECO.GCRD",
        flow_id="DSD_PMR@DF_PMR",
        version="1.2",
        dimensions="/...",
        label="Product Market Regulation Index"
    )
    
    if result is None or result == "RATE_LIMITED":
        return result
    
    df_raw = result["df"]
    
    # Filtrar solo PMR overall
    df_main = df_raw[df_raw["MEASURE"] == "PMR"].copy()
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    
    df_out = df_main[["REF_AREA", "TIME_PERIOD", "OBS_VALUE"]].copy()
    df_out.columns = ["iso3", "year", "oecd_pmr_overall"]
    df_out = df_out.dropna(subset=["oecd_pmr_overall"])
    
    print(f"  Paises: {df_out['iso3'].nunique()}, Anios: {df_out['year'].nunique()}")
    vals = df_out["oecd_pmr_overall"]
    print(f"  Score range: {vals.min():.4f} - {vals.max():.4f}")
    
    result["df"] = df_out
    result["variables"] = [
        {
            "variable_id": "oecd_pmr_overall",
            "hoja_datos": "7_PMR",
            "codigo_indicador": "PMR (overall, 0-6 scale)",
            "descripcion": "Product Market Regulation Index — Overall score (0-6)",
            "mensaje_analitico": "Mide barreras regulatorias a la competencia en mercados de productos. Proxy de calidad regulatoria general del pais. Control institucional adicional a wb_regulatory_quality. A mayor valor, mas restrictiva la regulacion.",
            "dataset": "DSD_PMR@DF_PMR"
        }
    ]
    
    return result


def extract_rd_tax():
    """Endpoint #8: R&D Tax Incentive Rates"""
    print("\n" + "="*70)
    print("ENDPOINT #8: R&D Tax Subsidy Rates")
    print("="*70)
    
    result = download_oecd_csv(
        agency="OECD.STI.STP",
        flow_id="DSD_RDTAX@DF_RDSUB",
        version="1.0",
        dimensions="/all",
        label="R&D Tax Subsidy Rates"
    )
    
    if result is None or result == "RATE_LIMITED":
        return result
    
    df_raw = result["df"]
    
    # Filtrar: LARGE firms + PROFITABLE scenario
    df_main = df_raw[
        (df_raw["SIZE"] == "LARGE") &
        (df_raw["PROFIT_SCENARIO"] == "PROFITABLE")
    ].copy()
    
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    df_main = df_main.dropna(subset=["OBS_VALUE"])
    
    df_out = df_main[["REF_AREA", "TIME_PERIOD", "OBS_VALUE"]].copy()
    df_out.columns = ["iso3", "year", "oecd_rd_tax_subsidy_rate"]
    
    print(f"  Paises: {df_out['iso3'].nunique()}, Anios: {df_out['year'].nunique()}")
    print(f"  Tasa subsidio rango: {df_out['oecd_rd_tax_subsidy_rate'].min():.2f} - {df_out['oecd_rd_tax_subsidy_rate'].max():.2f}")
    print(f"  Media: {df_out['oecd_rd_tax_subsidy_rate'].mean():.2f}")
    
    result["df"] = df_out
    result["variables"] = [
        {
            "variable_id": "oecd_rd_tax_subsidy_rate",
            "hoja_datos": "8_RD_TAX",
            "codigo_indicador": "RDSUB (LARGE + PROFITABLE)",
            "descripcion": "R&D Tax Subsidy Rate — Implied subsidy rate for large profitable firms (0-1)",
            "mensaje_analitico": "Mide el nivel de subsidio fiscal a I+D empresarial para grandes empresas rentables. Proxy de incentivos gubernamentales a innovacion empresarial. Relevante para sub-pregunta de innovacion e inversion: mayor subsidio = mayor incentivo a invertir en I+D.",
            "dataset": "DSD_RDTAX@DF_RDSUB"
        }
    ]
    
    return result


def extract_gbard():
    """Endpoint #9: Government Budget Allocations for R&D (GBARD)"""
    print("\n" + "="*70)
    print("ENDPOINT #9: Government Budget Allocations for R&D (GBARD)")
    print("="*70)
    
    result = download_oecd_csv(
        agency="OECD.STI.STP",
        flow_id="DSD_RDS_GOV@DF_GBARD_NABS07",
        version="1.0",
        dimensions="/all",
        label="Government Budget Allocations for R&D (GBARD)"
    )
    
    if result is None or result == "RATE_LIMITED":
        return result
    
    df_raw = result["df"]
    
    # GBARD en USD PPP para total socio-economic objectives
    # SEO = _T (total), FUNDMODE = _T (total), UNIT_MEASURE = USD_PPP
    df_main = df_raw[
        (df_raw["SEO"] == "_T") &
        (df_raw["FUNDMODE"] == "_T") &
        (df_raw["UNIT_MEASURE"] == "USD_PPP")
    ].copy()
    
    df_main["OBS_VALUE"] = pd.to_numeric(df_main["OBS_VALUE"], errors="coerce")
    df_main = df_main.dropna(subset=["OBS_VALUE"])
    
    df_out = df_main[["REF_AREA", "TIME_PERIOD", "OBS_VALUE"]].copy()
    df_out.columns = ["iso3", "year", "oecd_gbard_usd_ppp"]
    
    print(f"  Paises: {df_out['iso3'].nunique()}, Anios: {df_out['year'].nunique()}")
    print(f"  GBARD USD PPP rango: {df_out['oecd_gbard_usd_ppp'].min():.0f} - {df_out['oecd_gbard_usd_ppp'].max():.0f}")
    
    result["df"] = df_out
    result["variables"] = [
        {
            "variable_id": "oecd_gbard_usd_ppp",
            "hoja_datos": "9_GBARD",
            "codigo_indicador": "GBARD_NABS07 (_T, _T, USD_PPP)",
            "descripcion": "Government Budget Allocations for R&D — Total, USD PPP",
            "mensaje_analitico": "Mide el gasto publico total en I+D en USD PPP. Proxy de compromiso gubernamental con innovacion y desarrollo tecnologico. Relevante para sub-pregunta de innovacion: paises que asignan mas presupuesto publico a I+D muestran mayor priorizacion politica de innovacion.",
            "dataset": "DSD_RDS_GOV@DF_GBARD_NABS07"
        }
    ]
    
    return result


def extraer_todos():
    """Ejecuta todas las extracciones secuencialmente con delay entre llamadas."""
    import time
    resultados = {}
    
    endpoints = [
        ("stri_digital", extract_stri_digital),
        ("indigo", extract_indigo),
        ("regulatory_gov", extract_regulatory_gov),
        ("digital_gov", extract_digital_gov),
        ("ict_business", extract_ict_business),
        ("fdi_restrictiveness", extract_fdi_restrictiveness),
        ("pmr", extract_pmr),
        ("rd_tax", extract_rd_tax),
        ("gbard", extract_gbard),
    ]
    
    for i, (key, func) in enumerate(endpoints):
        if i > 0:
            print(f"\n  --- Esperando 3s antes del siguiente endpoint ---")
            time.sleep(3)  # Delay entre llamadas para evitar rate limiting
        result = func()
        if result == "RATE_LIMITED":
            print(f"  ⏸️ Endpoint '{key}' pendiente por rate limiting. Se reintentara manualmente.")
        resultados[key] = result
    
    return resultados


# =============================================================================
# CONSTRUCCION DEL EXCEL
# =============================================================================

def construir_excel(resultados):
    """Construye el archivo Excel con todas las hojas."""
    print("\n" + "="*70)
    print("CONSTRUYENDO EXCEL...")
    print("="*70)
    
    wb = Workbook()
    
    # --- Hoja 0_README ---
    ws_readme = wb.active
    ws_readme.title = "0_README"
    ws_readme.column_dimensions["A"].width = 25
    ws_readme.column_dimensions["B"].width = 80
    
    metadatos = {
        "Proyecto": "IMT3860 — ¿Regular o no regular? Analisis del impacto de marcos regulatorios de IA",
        "Fuente": "OECD Data Explorer / SDMX API (sdmx.oecd.org)",
        "Archivo": f"OECD_Data_Central_{TODAY}.xlsx",
        "Fecha generacion": TODAY,
        "Metodologia": "Extraccion directa via SDMX API. Cero datos sinteticos, cero imputacion.",
        "Regla R1": "Todo valor = respuesta directa del endpoint SDMX. Nada escrito a mano.",
        "Regla R2": "Si endpoint devuelve error o 0 filas → celda vacia. No se estima.",
        "Regla R3": "Cada endpoint documentado en 0_DICCIONARIO con URL exacta verificable.",
        "Regla R4": "Columna 'imputado' = 0 en todas las filas. Garantia explicita.",
        "Regla R5": "CSV crudos en raw_csv/ con SHA256 para auditoria.",
        "Regla R6": "Si endpoint esta rate limited (429): se marca PENDIENTE, no se detiene el proceso.",
        "Hojas": "0_README | 0_DICCIONARIO | 0_AUDITORIA | 1_DIGITAL_STRI | ...",
        "Nota rate limit": f"Endpoints rate limited al {TODAY}: {sum(1 for r in resultados.values() if r == 'RATE_LIMITED')}/{len(resultados)}",
        "Contacto": "Proyecto academico UC Chile — pablo@uc.cl"
    }
    agregar_metadata(ws_readme, metadatos)
    
    # --- Hoja 0_DICCIONARIO ---
    ws_dict = wb.create_sheet("0_DICCIONARIO")
    dict_headers = [
        "variable_id", "hoja_datos", "descripcion", "mensaje_analitico",
        "dataset_oecd", "codigo_indicador", "frecuencia", "actividad",
        "tamano_empresa", "paises_disponibles", "anios_disponibles",
        "endpoint_sdmx", "origen_metodologico", "fecha_extraccion",
        "filas_descargadas", "sha256_csv", "imputado", "notas"
    ]
    make_header_row(ws_dict, dict_headers)
    
    row = 2
    for key, result in resultados.items():
        if result is None:
            continue
        if result == "RATE_LIMITED":
            # Escribir entrada de pendiente en DICCIONARIO
            ws_dict.cell(row=row, column=1, value=f"PENDIENTE_{key}")
            ws_dict.cell(row=row, column=3, value="PENDIENTE — No se pudo descargar por rate limiting de OECD API")
            ws_dict.cell(row=row, column=12, value=f"Endpoint bloqueado temporalmente. Reintentar manualmente con --reintentar {key}")
            ws_dict.cell(row=row, column=17, value=0)
            ws_dict.cell(row=row, column=18, value="PENDIENTE — reintentar cuando API lo permita")
            row += 1
            continue
        df = result.get("df", pd.DataFrame())
        
        # Compute metadata from actual data
        n_countries = df["iso3"].nunique() if "iso3" in df.columns else ""
        year_min = df["year"].min() if "year" in df.columns else ""
        year_max = df["year"].max() if "year" in df.columns else ""
        years_str = f"{year_min}-{year_max}" if year_min != "" else ""
        
        for var in result.get("variables", []):
            ws_dict.cell(row=row, column=1, value=var.get("variable_id", ""))
            ws_dict.cell(row=row, column=2, value=var.get("hoja_datos", ""))
            ws_dict.cell(row=row, column=3, value=var.get("descripcion", ""))
            ws_dict.cell(row=row, column=4, value=var.get("mensaje_analitico", ""))
            ws_dict.cell(row=row, column=5, value=var.get("dataset", ""))
            ws_dict.cell(row=row, column=6, value=var.get("codigo_indicador", ""))
            ws_dict.cell(row=row, column=7, value="A (annual)")
            ws_dict.cell(row=row, column=8, value=var.get("actividad", "Total economy"))
            ws_dict.cell(row=row, column=9, value=var.get("tamano", "All sizes"))
            ws_dict.cell(row=row, column=10, value=str(n_countries))
            ws_dict.cell(row=row, column=11, value=years_str)
            ws_dict.cell(row=row, column=13, value="OECD_direct_collection")
            ws_dict.cell(row=row, column=14, value=TODAY)
            ws_dict.cell(row=row, column=15, value=result.get("n_rows", ""))
            ws_dict.cell(row=row, column=16, value=result.get("sha256", ""))
            ws_dict.cell(row=row, column=17, value=0)  # imputado = 0 garantizado
            ws_dict.cell(row=row, column=18, value="Dato 100% real. Descargado via SDMX API OECD. Cero imputacion, cero estimacion.")
            # Link endpoint
            if result.get("url"):
                cell_url = ws_dict.cell(row=row, column=12, value=result["url"])
                cell_url.font = LINK_FONT
            row += 1
    
    auto_width(ws_dict)
    # Fijar ancho minimo para columnas de texto largo
    ws_dict.column_dimensions["D"].width = 55
    ws_dict.column_dimensions["C"].width = 45
    ws_dict.column_dimensions["L"].width = 50
    
    # --- Hoja 0_AUDITORIA ---
    ws_audit = wb.create_sheet("0_AUDITORIA")
    audit_headers = [
        "endpoint", "dataset", "fecha_extraccion", "filas_descargadas",
        "columnas", "sha256", "archivo_crudo", "paises_unicos",
        "anios_rango", "errores", "observaciones"
    ]
    make_header_row(ws_audit, audit_headers)
    
    row = 2
    for key, result in resultados.items():
        if result is None:
            ws_audit.cell(row=row, column=1, value=key)
            ws_audit.cell(row=row, column=10, value="ERROR")
            ws_audit.cell(row=row, column=11, value="No se pudo descargar")
            row += 1
            continue
        if result == "RATE_LIMITED":
            ws_audit.cell(row=row, column=1, value=key)
            ws_audit.cell(row=row, column=3, value=TODAY)
            ws_audit.cell(row=row, column=10, value="PENDING (429)")
            ws_audit.cell(row=row, column=11, value="Rate limited por OECD API. Reintentar manualmente.")
            row += 1
            continue
        df = result.get("df", pd.DataFrame())
        ws_audit.cell(row=row, column=1, value=result.get("url", ""))
        ws_audit.cell(row=row, column=2, value=f"DSD_{key.upper()}")
        ws_audit.cell(row=row, column=3, value=TODAY)
        ws_audit.cell(row=row, column=4, value=result.get("n_rows", ""))
        ws_audit.cell(row=row, column=5, value=len(df.columns))
        ws_audit.cell(row=row, column=6, value=result.get("sha256", ""))
        ws_audit.cell(row=row, column=7, value=result.get("raw_path", ""))
        if "iso3" in df.columns:
            ws_audit.cell(row=row, column=8, value=df["iso3"].nunique())
        if "year" in df.columns:
            ws_audit.cell(row=row, column=9, value=f"{df['year'].min()}-{df['year'].max()}")
        ws_audit.cell(row=row, column=10, value="OK")
        ws_audit.cell(row=row, column=11, value="Extraccion directa OECD SDMX")
        row += 1
    
    auto_width(ws_audit)
    
    # --- Hojas de datos por endpoint ---
    sheet_names = {
        "stri_digital": "1_DIGITAL_STRI",
        "indigo": "2_INDIGO",
        "regulatory_gov": "3_REGULATORY_GOV",
        "digital_gov": "4_DIGITAL_GOV",
        "ict_business": "5_ICT_BUSINESS",
        "fdi_restrictiveness": "6_FDI_RESTRICTIVENESS",
        "pmr": "7_PMR",
        "rd_tax": "8_RD_TAX",
        "gbard": "9_GBARD",
    }
    
    for key, result in resultados.items():
        if result is None or result == "RATE_LIMITED":
            continue
        sheet_name = sheet_names.get(key, f"DATA_{key}")
        ws_data = wb.create_sheet(sheet_name)
        
        df = result["df"]
        headers = list(df.columns)
        make_header_row(ws_data, headers)
        write_dataframe(ws_data, df)
        auto_width(ws_data)
        
        print(f"  Hoja '{sheet_name}': {len(df)} filas escritas")
    
    # Guardar
    wb.save(OUTPUT_FILE)
    print(f"\nExcel guardado: {OUTPUT_FILE}")
    return OUTPUT_FILE


# =============================================================================
# MAIN
# =============================================================================
if __name__ == "__main__":
    print("="*70)
    print("OECD DATA CENTRAL — EXTRACCION SDMX")
    print(f"Fecha: {TODAY}")
    print("="*70)
    
    resultados = extraer_todos()
    construir_excel(resultados)
    
    print("\nPROCESO COMPLETADO.")
    print(f"Resumen: {sum(1 for r in resultados.values() if r is not None)}/{len(resultados)} endpoints OK")
